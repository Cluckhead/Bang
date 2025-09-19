# Purpose: Export each Excel sheet to two CSVs (values and formulas) for LLM-friendly analysis.
"""
This script reads an Excel workbook and, for every sheet, writes:
- A values CSV containing the evaluated cell values
- A formulas CSV containing only the cell formulas (blank where cells are constants)

Intended to make spreadsheets easier for LLMs to consume by separating computed values from logic.
"""

from __future__ import annotations

import argparse
import csv
import datetime as _dt
import json
import os
import re
import sys
from pathlib import Path
from typing import Iterable, List, Optional


def _ensure_openpyxl_available() -> None:
    try:
        import openpyxl  # noqa: F401
    except Exception as exc:  # pragma: no cover - environment specific
        message = (
            "Required dependency 'openpyxl' is not installed.\n"
            "Install it with PowerShell:\n\n"
            "    python -m pip install --upgrade openpyxl\n\n"
            f"Original error: {exc}"
        )
        print(message, file=sys.stderr)
        sys.exit(2)


def _sanitize_sheet_name(sheet_name: str) -> str:
    """Create a filesystem-safe version of the sheet name for filenames."""
    # Replace invalid filename characters and trim length
    safe = re.sub(r"[^A-Za-z0-9._-]+", "_", sheet_name.strip())
    # Avoid reserved names on Windows (CON, PRN, AUX, NUL, COM1.., LPT1..)
    reserved = {"CON", "PRN", "AUX", "NUL"} | {f"COM{i}" for i in range(1, 10)} | {f"LPT{i}" for i in range(1, 10)}
    if safe.upper() in reserved:
        safe = f"_{safe}"
    return safe[:80] if len(safe) > 80 else safe


def _normalize_value(value: object) -> str:
    """Normalize Python/Excel values into a CSV-friendly string."""
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        # Represent numbers plainly; csv module will handle quoting
        return str(value)
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, _dt.datetime):
        # Preserve timezone if present; Excel datetimes are typically naive
        return value.isoformat(sep=" ")
    if isinstance(value, _dt.date):
        return value.isoformat()
    if isinstance(value, _dt.time):
        return value.isoformat()
    # Fallback to string
    return str(value)


def _write_csv(output_path: Path, rows: Iterable[Iterable[object]]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for row in rows:
            writer.writerow([_normalize_value(cell) for cell in row])


def _sheet_dimensions(ws) -> tuple[int, int]:
    # ws.max_row / max_column give upper bounds, even if trailing rows are empty
    # We keep them to preserve layout; trimming would diverge values/formulas shapes.
    return ws.max_row or 0, ws.max_column or 0


def _iter_values_grid(ws) -> Iterable[List[object]]:
    for row in ws.iter_rows(values_only=True):
        yield list(row)


def _iter_formulas_grid(ws) -> Iterable[List[object]]:
    # When data_only=False, cell.value is the formula string (e.g., "=SUM(A1:A3)")
    # for formula cells; for constants, it is the constant. We want blanks for constants.
    for row in ws.iter_rows(values_only=False):
        out_row: List[object] = []
        for cell in row:
            value = cell.value
            is_formula = False
            # Prefer data_type check when available
            try:
                is_formula = getattr(cell, "data_type", None) == "f"
            except Exception:
                is_formula = False
            if not is_formula and isinstance(value, str) and value.startswith("="):
                is_formula = True
            out_row.append(value if is_formula else "")
        yield out_row


def export_excel_to_csvs(
    excel_path: Path,
    output_dir: Path,
    include_hidden_sheets: bool = False,
    selected_sheets: Optional[List[str]] = None,
) -> dict:
    """Export each sheet to two CSVs (values and formulas). Returns summary metadata."""
    _ensure_openpyxl_available()
    import openpyxl  # Imported after check

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    # Open twice: one for evaluated values, one for formulas
    wb_values = openpyxl.load_workbook(filename=str(excel_path), data_only=True, read_only=True)
    wb_formulas = openpyxl.load_workbook(filename=str(excel_path), data_only=False, read_only=True)

    sheet_names = [s for s in wb_values.sheetnames if s in wb_formulas.sheetnames]

    if selected_sheets:
        selected_set = {s.strip() for s in selected_sheets}
        sheet_names = [s for s in sheet_names if s in selected_set]

    summary = {
        "source": str(excel_path),
        "output_dir": str(output_dir),
        "sheets": [],
    }

    for sheet_name in sheet_names:
        ws_values = wb_values[sheet_name]
        ws_formulas = wb_formulas[sheet_name]

        if not include_hidden_sheets and (ws_values.sheet_state != "visible"):
            continue

        safe_name = _sanitize_sheet_name(sheet_name)

        values_out = output_dir / f"{safe_name}.values.csv"
        formulas_out = output_dir / f"{safe_name}.formulas.csv"

        _write_csv(values_out, _iter_values_grid(ws_values))
        _write_csv(formulas_out, _iter_formulas_grid(ws_formulas))

        summary["sheets"].append(
            {
                "sheet": sheet_name,
                "state": ws_values.sheet_state,
                "values_csv": str(values_out),
                "formulas_csv": str(formulas_out),
            }
        )

    # Optionally write an index file to help discovery
    index_path = output_dir / "index.json"
    index_path.parent.mkdir(parents=True, exist_ok=True)
    index_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")

    return summary


def _build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Export each Excel sheet to two CSV files: one for values and one for formulas."
        )
    )
    parser.add_argument(
        "-i",
        "--input",
        required=True,
        help="Path to the Excel file (.xlsx/.xlsm)",
    )
    parser.add_argument(
        "-o",
        "--output",
        required=False,
        help=(
            "Directory to write CSVs. Defaults to '<excel_basename>_csvs' next to the input file."
        ),
    )
    parser.add_argument(
        "-s",
        "--sheets",
        required=False,
        help="Comma-separated list of sheet names to export (default: all visible)",
    )
    parser.add_argument(
        "--include-hidden",
        action="store_true",
        help="Include hidden sheets",
    )
    return parser


def main(argv: Optional[List[str]] = None) -> int:
    parser = _build_arg_parser()
    args = parser.parse_args(argv)

    excel_path = Path(args.input).expanduser().resolve()
    if args.output:
        output_dir = Path(args.output).expanduser().resolve()
    else:
        output_dir = excel_path.with_name(f"{excel_path.stem}_csvs")

    selected_sheets = [s.strip() for s in args.sheets.split(",")] if args.sheets else None

    try:
        summary = export_excel_to_csvs(
            excel_path=excel_path,
            output_dir=output_dir,
            include_hidden_sheets=args.include_hidden,
            selected_sheets=selected_sheets,
        )
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1

    # Print a concise summary to stdout
    print(json.dumps(summary, indent=2))
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())

