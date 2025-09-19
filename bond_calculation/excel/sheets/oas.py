# sheets/oas.py
# Purpose: Build the OAS_Calculation sheet

from __future__ import annotations

from typing import Dict
from openpyxl import Workbook
from openpyxl.styles import Font


def add_oas_sheet(wb: Workbook, bond_data: Dict, python_results: Dict) -> None:
    ws = wb.create_sheet("OAS_Calculation")
    ws.append(["OPTION-ADJUSTED SPREAD (OAS) CALCULATION"])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])

    ws.append(["What is OAS?"])
    ws['A3'].font = Font(bold=True, size=12, color="0066CC")
    ws.append(["OAS is the spread after removing the value of embedded options"])
    ws.append(["Formula: OAS = Z-Spread - Option Value (in spread terms)"])
    ws.append([])

    ws.append(["Standard OAS Calculation (Single Call, Fixed Volatility)"])
    ws['A7'].font = Font(bold=True, size=12, color="0066CC")

    if bond_data.get('call_schedule') and python_results.get('oas_standard') is not None:
        ws.append(["Method:", "Black Model"])
        ws.append(["Volatility:", f"{python_results['oas_details'].get('standard_volatility', 0.20)*100:.0f}%"])
        ws.append(["Calls Used:", f"{python_results['oas_details'].get('calls_used', 1)} of {len(bond_data.get('call_schedule') or [])}"])
        ws.append(["Z-Spread:", f"{python_results['z_spread']*10000:.1f} bps"])
        ws.append(["OAS Result:", f"{python_results['oas_standard']*10000:.1f} bps"])
        ws.append(["Option Value:", f"{(python_results['z_spread'] - python_results['oas_standard'])*10000:.1f} bps"])
    else:
        ws.append(["Not applicable: no embedded options or calculation failed"])

    ws.append([])
    ws.append(["Enhanced OAS Calculation (All Calls, Market Volatility)"])
    ws['A15'].font = Font(bold=True, size=12, color="0066CC")

    if bond_data.get('call_schedule') and python_results.get('oas_enhanced') is not None:
        ws.append(["Method:", python_results['oas_details'].get('method', 'Binomial Tree')])
        ws.append(["Volatility:", f"{python_results['oas_details'].get('enhanced_volatility', 0.15)*100:.1f}% (calibrated)"])
        ws.append(["Calls Used:", f"All {len(bond_data.get('call_schedule') or [])}"])
        ws.append(["Z-Spread:", f"{python_results['z_spread']*10000:.1f} bps"])
        ws.append(["OAS Result:", f"{python_results['oas_enhanced']*10000:.1f} bps"])
        ws.append(["Option Value:", f"{(python_results['z_spread'] - python_results['oas_enhanced'])*10000:.1f} bps"])
        if python_results.get('oas_standard') is not None:
            ws.append([])
            ws.append(["Improvement over Standard:"])
            improvement = (python_results['oas_enhanced'] - python_results['oas_standard']) * 10000
            ws.append([f"  Difference: {improvement:+.1f} bps"])
            ws.append(["  • Uses market-calibrated volatility"])
            ws.append(["  • Considers all call dates"])
            ws.append(["  • American option valuation"])
    else:
        ws.append(["Not applicable: enhanced OAS requires callable features"])


