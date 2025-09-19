# sheets/data_source.py
# Purpose: Build the Data_Source sheet with raw data blocks for auditability

from __future__ import annotations

from datetime import datetime
from typing import Dict, List, Tuple
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter


def add_data_source_sheet(
    wb: Workbook,
    bond_data: Dict,
    cashflows: List[Dict],
    curve_data: Tuple[List[float], List[float]],
    price: float,
    valuation_date: datetime,
    sec_data: Dict = None,
    accrued_interest: float = None
) -> None:
    """
    Create a Data_Source sheet with raw data blocks used by the workbook.
    This provides transparency and auditability for all source data.
    """
    ws = wb.create_sheet("Data_Source")
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "RAW DATA SOURCE - AUDIT TRAIL"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Section styling
    section_font = Font(bold=True, size=11, color="FFFFFF")
    section_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D5DBDB", end_color="D5DBDB", fill_type="solid")
    
    row = 3
    
    # Reference Data Section
    ws[f'A{row}'] = "REFERENCE DATA"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    
    # Add reference data headers
    ref_headers = ["Field", "Value", "Source"]
    for col, header in enumerate(ref_headers, 1):
        ws.cell(row, col, header)
        ws.cell(row, col).font = header_font
        ws.cell(row, col).fill = header_fill
    row += 1
    
    # Add reference data
    ref_start_row = row
    for key, value in bond_data['reference'].items():
        ws[f'A{row}'] = key
        if isinstance(value, datetime):
            ws[f'B{row}'] = value.strftime('%d/%m/%Y')
        else:
            ws[f'B{row}'] = value
        ws[f'C{row}'] = "reference.csv"
        row += 1
    
    # Define named range for reference data
    ref_end_row = row - 1
    wb.defined_names.add(DefinedName("Data_Reference", 
                                     attr_text=f"Data_Source!$A${ref_start_row}:$C${ref_end_row}"))
    row += 2
    
    # Schedule Data Section
    ws[f'A{row}'] = "SCHEDULE DATA"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    
    # Add schedule headers
    schedule_headers = ["Field", "Value", "Source"]
    for col, header in enumerate(schedule_headers, 1):
        ws.cell(row, col, header)
        ws.cell(row, col).font = header_font
        ws.cell(row, col).fill = header_fill
    row += 1
    
    # Add schedule data
    schedule_start_row = row
    for key, value in bond_data['schedule'].items():
        ws[f'A{row}'] = key
        if isinstance(value, datetime):
            ws[f'B{row}'] = value.strftime('%d/%m/%Y')
        elif isinstance(value, list) and len(value) > 0 and isinstance(value[0], datetime):
            ws[f'B{row}'] = ', '.join([d.strftime('%d/%m/%Y') for d in value[:5]]) + ('...' if len(value) > 5 else '')
        else:
            ws[f'B{row}'] = str(value)
        ws[f'C{row}'] = "schedule.csv"
        row += 1
    
    schedule_end_row = row - 1
    wb.defined_names.add(DefinedName("Data_Schedule", 
                                     attr_text=f"Data_Source!$A${schedule_start_row}:$C${schedule_end_row}"))
    row += 2
    
    # Yield Curve Data Section
    ws[f'A{row}'] = "YIELD CURVE DATA"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    
    # Add curve headers
    curve_headers = ["Tenor (Years)", "Zero Rate (%)", "Source", "Val Date", "Currency"]
    for col, header in enumerate(curve_headers, 1):
        ws.cell(row, col, header)
        ws.cell(row, col).font = header_font
        ws.cell(row, col).fill = header_fill
    row += 1
    
    # Add curve data
    curve_start_row = row
    terms, rates = curve_data
    for term, rate in zip(terms, rates):
        ws[f'A{row}'] = term
        ws[f'B{row}'] = rate
        ws[f'C{row}'] = "curves.csv"
        ws[f'D{row}'] = valuation_date.strftime('%d/%m/%Y')
        ws[f'E{row}'] = bond_data['reference'].get('Position Currency', 'USD')
        row += 1
    
    curve_end_row = row - 1
    wb.defined_names.add(DefinedName("Data_Curve", 
                                     attr_text=f"Data_Source!$A${curve_start_row}:$E${curve_end_row}"))
    
    # Define specific named ranges for curve terms and rates (for formulas)
    wb.defined_names.add(DefinedName("Curve_Terms", 
                                     attr_text=f"Data_Source!$A${curve_start_row}:$A${curve_end_row}"))
    wb.defined_names.add(DefinedName("Curve_Rates", 
                                     attr_text=f"Data_Source!$B${curve_start_row}:$B${curve_end_row}"))
    row += 2
    
    # Cashflow Data Section
    ws[f'A{row}'] = "CASHFLOW SCHEDULE"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    
    # Add cashflow headers
    cf_headers = ["Date", "Type", "Amount", "Days from Val", "Accrual Fraction", "Source"]
    for col, header in enumerate(cf_headers, 1):
        ws.cell(row, col, header)
        ws.cell(row, col).font = header_font
        ws.cell(row, col).fill = header_fill
    row += 1
    
    # Add cashflow data
    cf_start_row = row
    for cf in cashflows:
        # Handle date formatting
        cf_date = cf.get('date')
        if hasattr(cf_date, 'strftime'):
            date_str = cf_date.strftime('%d/%m/%Y')
        else:
            date_str = cf_date
        
        # Determine payment type
        coupon = cf.get('coupon', 0)
        principal = cf.get('principal', 0)
        if principal > 0 and coupon > 0:
            cf_type = 'Principal+Coupon'
        elif principal > 0:
            cf_type = 'Principal'
        else:
            cf_type = 'Coupon'
        
        # Get amount (could be 'amount' or 'total')
        amount = cf.get('amount', cf.get('total', coupon + principal))
        
        ws[f'A{row}'] = date_str
        ws[f'B{row}'] = cf_type
        ws[f'C{row}'] = amount
        ws[f'D{row}'] = cf.get('days_from_valuation', '')
        ws[f'E{row}'] = cf.get('accrual_period', cf.get('accrual_fraction', ''))
        ws[f'F{row}'] = "Python calculation"
        row += 1
    
    cf_end_row = row - 1
    wb.defined_names.add(DefinedName("Data_Cashflows", 
                                     attr_text=f"Data_Source!$A${cf_start_row}:$F${cf_end_row}"))
    row += 2
    
    # Market Data Section
    ws[f'A{row}'] = "MARKET DATA"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    
    # Add market data
    market_headers = ["Field", "Value", "Date", "Source"]
    for col, header in enumerate(market_headers, 1):
        ws.cell(row, col, header)
        ws.cell(row, col).font = header_font
        ws.cell(row, col).fill = header_fill
    row += 1
    
    # Use file-based accrued if available
    if accrued_interest is not None:
        accrued_value = accrued_interest
        accrued_source = "sec_accrued.csv"
    else:
        accrued_value = 0.0
        accrued_source = "Default (file not found)"
    
    market_data = [
        ("Clean Price", price, valuation_date.strftime('%d/%m/%Y'), "Market/User Input"),
        ("Accrued Interest (File)", accrued_value, valuation_date.strftime('%d/%m/%Y'), accrued_source),
        ("Accrued (Excel Formula)", "=ACCRINT(COUPPCD(Sel_ValDate,Maturity,Sel_Freq,Assump_Basis_Code),COUPNCD(Sel_ValDate,Maturity,Sel_Freq,Assump_Basis_Code),Sel_ValDate,Coupon_Rate/100,100,Sel_Freq,Assump_Basis_Code,TRUE)", 
         valuation_date.strftime('%d/%m/%Y'), "Excel Calculation (reference)"),
        ("Dirty Price", f"={price}+{accrued_value}", valuation_date.strftime('%d/%m/%Y'), "Clean + File Accrued"),
    ]
    
    for field, value, date, source in market_data:
        ws[f'A{row}'] = field
        ws[f'B{row}'] = value
        ws[f'C{row}'] = date
        ws[f'D{row}'] = source
        row += 1
    
    # Define named ranges for pricing (point to file-based accrued, not Excel formula)
    # The rows were: Clean Price, Accrued (File), Accrued (Excel), Dirty
    # So: row-4 = Clean, row-3 = Accrued File, row-2 = Accrued Excel, row-1 = Dirty
    wb.defined_names.add(DefinedName("Price_Accrued", attr_text=f"Data_Source!$B${row-3}"))  # Points to file-based accrued
    wb.defined_names.add(DefinedName("Price_Dirty", attr_text=f"Data_Source!$B${row-1}"))
    row += 2
    
    # Security Data Section (if available)
    if sec_data:
        ws[f'A{row}'] = "SECURITY TIME SERIES DATA"
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        ws.merge_cells(f'A{row}:H{row}')
        row += 1
        
        # Add security data headers
        sec_headers = ["Date", "Price", "Spread", "YTM", "Duration", "Source"]
        for col, header in enumerate(sec_headers, 1):
            ws.cell(row, col, header)
            ws.cell(row, col).font = header_font
            ws.cell(row, col).fill = header_fill
        row += 1
        
        # Add sample security data (would be populated from sec_*.csv files)
        sec_start_row = row
        if 'time_series' in sec_data:
            for ts_point in sec_data['time_series'][-10:]:  # Last 10 points
                ws[f'A{row}'] = ts_point.get('date', '')
                ws[f'B{row}'] = ts_point.get('price', '')
                ws[f'C{row}'] = ts_point.get('spread', '')
                ws[f'D{row}'] = ts_point.get('ytm', '')
                ws[f'E{row}'] = ts_point.get('duration', '')
                ws[f'F{row}'] = "sec_*.csv"
                row += 1
        
        if row > sec_start_row:
            sec_end_row = row - 1
            wb.defined_names.add(DefinedName("Data_SecTimeSeries", 
                                             attr_text=f"Data_Source!$A${sec_start_row}:$F${sec_end_row}"))
    
    # Metadata Section
    row += 2
    ws[f'A{row}'] = "METADATA"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    
    metadata = [
        ("Data Extraction Time", datetime.now().strftime('%d/%m/%Y %H:%M:%S')),
        ("Python Version", "SpreadOMatic v1.0"),
        ("Excel Generation", "bond_calculation.excel.workbook"),
        ("Data Sources", "Data/*.csv files"),
        ("Valuation Date", valuation_date.strftime('%d/%m/%Y')),
        ("ISIN", bond_data['reference']['ISIN']),
    ]
    
    for label, value in metadata:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    # Freeze panes
    ws.freeze_panes = 'A4'