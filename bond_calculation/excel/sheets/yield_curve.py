# sheets/yield_curve.py
# Purpose: Build the Yield_Curve sheet (editable)

from __future__ import annotations

from typing import List, Tuple
from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName

from ..styles import header_font, header_fill, input_fill


def add_yield_curve_sheet(wb: Workbook, curve_data: Tuple[List[float], List[float]]) -> None:
    ws = wb.create_sheet("Yield_Curve")
    ws.append(["YIELD CURVE DATA (Editable)"])
    from openpyxl.styles import Font
    ws['A1'].font = Font(bold=True, size=12)
    ws.append([])
    ws.append(["Term (Years)", "Zero Rate (%)", "Discount Factor Formula"])

    for i in range(1, 4):
        cell = ws.cell(row=3, column=i)
        cell.font = header_font
        cell.fill = header_fill

    row = 4
    for t, r in zip(curve_data[0], curve_data[1]):
        ws.append([t, r * 100, f"=1/(1+B{row}/100)^A{row}"])
        ws.cell(row=row, column=2).fill = input_fill
        row += 1

    # Define named ranges used elsewhere
    last = row - 1
    def _add_defined(name: str, ref: str) -> None:
        try:
            dn = DefinedName(name=name, attr_text=ref)
            if hasattr(wb.defined_names, "add"):
                wb.defined_names.add(dn)
            elif hasattr(wb.defined_names, "append"):
                wb.defined_names.append(dn)
        except Exception:
            pass

    if last >= 4:
        _add_defined('Curve_Terms', f"Yield_Curve!$A$4:$A${last}")
        _add_defined('Curve_Rates', f"Yield_Curve!$B$4:$B${last}")
        _add_defined('Curve_DFs', f"Yield_Curve!$C$4:$C${last}")


