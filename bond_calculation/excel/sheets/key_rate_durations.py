# sheets/key_rate_durations.py
# Purpose: Build the Key_Rate_Durations sheet using python_results

from __future__ import annotations

from typing import Dict
from openpyxl import Workbook
from openpyxl.styles import Font
from ..styles import header_font, header_fill


def add_key_rate_durations_sheet(wb: Workbook, python_results: Dict) -> None:
    ws = wb.create_sheet("Key_Rate_Durations")
    ws.append(["KEY RATE DURATIONS (KRD)"])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Python Calculated KRDs:"])
    ws['A3'].font = Font(bold=True, color="0066CC")
    ws.append(["Tenor", "Duration", "Description"])
    for i in range(1, 4):
        cell = ws.cell(row=5, column=i)
        cell.font = header_font
        cell.fill = header_fill

    # Fill rows from results in canonical tenor order
    order = ["1Y", "2Y", "3Y", "5Y", "7Y", "10Y", "20Y", "30Y"]
    krds = python_results.get('key_rate_durations', {}) or {}
    row = 6
    for tenor in order:
        value = krds.get(tenor, 0.0)
        if isinstance(tenor, (int, float)):
            if float(tenor) < 1:
                desc = f"Sensitivity to {float(tenor)*12:.0f}M rate"
            else:
                desc = f"Sensitivity to {float(tenor):.1f}Y rate"
        else:
            desc = f"Sensitivity to {tenor} rate"
        ws.append([tenor, value, desc])
        ws.cell(row=row, column=2).number_format = '0.0000'
        row += 1

    ws.append([])
    ws.append(["Total KRD:", f"=SUM(B6:B{row-1})", "Should ≈ Effective Duration"])
    ws.cell(row=ws.max_row, column=2).number_format = '0.0000'


