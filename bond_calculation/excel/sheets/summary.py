# sheets/summary.py
# Purpose: Build the Summary_Comparison sheet

from __future__ import annotations

from typing import Dict

from openpyxl import Workbook
from ..styles import header_font, header_fill


def add_summary_sheet(wb: Workbook, bond_data: Dict, python_results: Dict) -> None:
    ws = wb.create_sheet("Summary_Comparison")
    ws.append(["CALCULATION METHODS COMPARISON"])
    from openpyxl.styles import Font
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])

    ws.append(["Spread Metrics", "Value", "Method", "Notes"])
    for i in range(1, 5):
        cell = ws.cell(row=3, column=i)
        cell.font = header_font
        cell.fill = header_fill

    ws.append(["YTM", f"{python_results['ytm']}", "SpreadOMatic Newton-Raphson", "Most accurate"])
    ws.append(["Z-Spread (bps)", f"{python_results['z_spread']*10000}", "SpreadOMatic calculation", "Calculated"])
    ws.append(["G-Spread (bps)", f"{python_results['g_spread']*10000}", "SpreadOMatic calculation", "YTM minus govt rate"])

    if bond_data.get('call_schedule'):
        ws.append(["OAS Standard (bps)", f"{python_results['oas_standard']*10000:.1f}" if python_results['oas_standard'] else "N/A", "Black Model, 20% vol", "Single call option"])
        ws.append(["OAS Enhanced (bps)", f"{python_results['oas_enhanced']*10000:.1f}" if (python_results.get('oas_enhanced') is not None and python_results.get('oas_enhanced') != "N/A") else "N/A", "Binomial Tree", "All calls, calibrated vol"])


