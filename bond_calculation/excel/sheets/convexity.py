# sheets/convexity.py
# Purpose: Build the Convexity sheet with python result and illustrative formulas

from __future__ import annotations

from typing import Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font
from ..styles import header_font, header_fill


def add_convexity_sheet(wb: Workbook, python_results: Dict, cashflows: List[Dict]) -> None:
    ws = wb.create_sheet("Convexity")
    ws.append(["CONVEXITY CALCULATION"])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Python Calculated Convexity:", python_results.get('convexity', 0.0)])
    ws['A3'].font = Font(bold=True, color="0066CC")
    ws.cell(row=3, column=2).number_format = '0.00'
    ws.append([])

    ws.append(["Price approximation table (illustrative):"])
    headers = ["Rate Change", "Actual Price", "Duration Approx", "Duration+Convexity"]
    ws.append(headers)
    for i in range(1, len(headers) + 1):
        cell = ws.cell(row=6, column=i)
        cell.font = header_font
        cell.fill = header_fill

    rate_changes = [-200, -100, -50, 0, 50, 100, 200]
    for rc in rate_changes:
        ws.append([
            rc,
            f"=SUMPRODUCT(Cashflows!I$2:I${len(cashflows)+1},1/(1+(FORECAST(Cashflows!D$2:D${len(cashflows)+1},Curve_Rates,Curve_Terms)+{rc}/10000)/Assump_Frequency)^(Assump_Frequency*Cashflows!D$2:D${len(cashflows)+1}))",
            f"=Price_Dirty*(1-Effective_Duration!B13*{rc}/10000)",
            f"=Price_Dirty*(1-Effective_Duration!B13*{rc}/10000+0.5*{python_results.get('convexity', 0.0)}*({rc}/10000)^2)",
        ])


