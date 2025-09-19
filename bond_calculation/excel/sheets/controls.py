# sheets/controls.py
# Purpose: Build the Controls sheet with dropdowns for interactive parameter selection

from __future__ import annotations

from datetime import datetime
from typing import Dict, List, Tuple
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter

from ..styles import header_font, header_fill, input_fill, highlight_fill


def add_controls_sheet(
    wb: Workbook, 
    bond_data: Dict, 
    price: float, 
    valuation_date: datetime
) -> None:
    """
    Create a Controls sheet with calculation parameters for a single bond.
    All selections will be defined as named ranges for use in formulas.
    """
    ws = wb.create_sheet("Controls", 0)  # Insert at beginning
    
    # Title
    ws.merge_cells('A1:E1')
    ws['A1'] = "BOND ANALYTICS CONTROL PANEL"
    ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Section headers styling
    section_font = Font(bold=True, size=12)
    section_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    section_font_white = Font(bold=True, size=12, color="FFFFFF")
    
    # Input styling
    input_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row = 3
    
    # Bond Information Section (Read-only)
    ws[f'A{row}'] = "BOND INFORMATION"
    ws[f'A{row}'].font = section_font_white
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:C{row}')
    row += 1
    
    # ISIN (Read-only)
    ws[f'A{row}'] = "ISIN:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = bond_data['reference']['ISIN']
    ws[f'B{row}'].fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
    ws[f'B{row}'].border = input_border
    ws[f'C{row}'] = "Bond identifier"
    wb.defined_names.add(DefinedName("Sel_ISIN", attr_text=f"Controls!$B${row}"))
    row += 1
    
    # Security Name
    ws[f'A{row}'] = "Security Name:"
    ws[f'B{row}'] = bond_data['reference']['Security Name']
    ws[f'B{row}'].fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
    ws[f'B{row}'].border = input_border
    row += 2
    
    # Valuation Parameters Section
    ws[f'A{row}'] = "VALUATION PARAMETERS"
    ws[f'A{row}'].font = section_font_white
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:C{row}')
    row += 1
    
    # Valuation Date (Read-only for this bond/date)
    ws[f'A{row}'] = "Valuation Date:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = valuation_date.strftime('%d/%m/%Y')
    ws[f'B{row}'].fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
    ws[f'B{row}'].border = input_border
    ws[f'C{row}'] = "DD/MM/YYYY format"
    wb.defined_names.add(DefinedName("Sel_ValDate", attr_text=f"Controls!$B${row}"))
    row += 1
    
    # Currency (Read-only)
    ws[f'A{row}'] = "Currency:"
    ws[f'A{row}'].font = Font(bold=True)
    currency = bond_data['reference'].get('Position Currency', 'USD')
    ws[f'B{row}'] = currency
    ws[f'B{row}'].fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
    ws[f'B{row}'].border = input_border
    ws[f'C{row}'] = "Curve currency"
    wb.defined_names.add(DefinedName("Sel_Currency", attr_text=f"Controls!$B${row}"))
    row += 2
    
    # Pricing Section
    ws[f'A{row}'] = "PRICING"
    ws[f'A{row}'].font = section_font_white
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:C{row}')
    row += 1
    
    # Clean Price
    ws[f'A{row}'] = "Clean Price:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = price
    ws[f'B{row}'].fill = highlight_fill  # Different color to show it's editable
    ws[f'B{row}'].border = input_border
    ws[f'C{row}'] = "Market price (editable)"
    wb.defined_names.add(DefinedName("Price_Clean", attr_text=f"Controls!$B${row}"))
    row += 2
    
    # Calculation Parameters Section
    ws[f'A{row}'] = "CALCULATION PARAMETERS"
    ws[f'A{row}'].font = section_font_white
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:C{row}')
    row += 1
    
    # Coupon Frequency
    ws[f'A{row}'] = "Coupon Frequency:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = int(bond_data['schedule']['Coupon Frequency'])
    ws[f'B{row}'].fill = input_fill
    ws[f'B{row}'].border = input_border
    ws[f'C{row}'] = "Payments per year"
    dv_freq = DataValidation(type="list", formula1='"1,2,4,12"', allow_blank=False)
    dv_freq.add(f'B{row}')
    ws.add_data_validation(dv_freq)
    wb.defined_names.add(DefinedName("Sel_Freq", attr_text=f"Controls!$B${row}"))
    row += 1
    
    # Day Basis
    ws[f'A{row}'] = "Day Count Basis:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = bond_data['schedule']['Day Basis']
    ws[f'B{row}'].fill = input_fill
    ws[f'B{row}'].border = input_border
    ws[f'C{row}'] = "Day count convention"
    dv_basis = DataValidation(type="list", formula1='"ACT/ACT,30/360,30E/360,ACT/360,ACT/365"', allow_blank=False)
    dv_basis.add(f'B{row}')
    ws.add_data_validation(dv_basis)
    wb.defined_names.add(DefinedName("Sel_DayBasis", attr_text=f"Controls!$B${row}"))
    row += 1
    
    # Compounding
    ws[f'A{row}'] = "Compounding:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = "Semiannual"
    ws[f'B{row}'].fill = input_fill
    ws[f'B{row}'].border = input_border
    ws[f'C{row}'] = "Yield compounding"
    dv_comp = DataValidation(type="list", formula1='"Annual,Semiannual,Quarterly,Monthly,Continuous"', allow_blank=False)
    dv_comp.add(f'B{row}')
    ws.add_data_validation(dv_comp)
    wb.defined_names.add(DefinedName("Sel_Compounding", attr_text=f"Controls!$B${row}"))
    row += 1
    
    # Settlement Convention
    ws[f'A{row}'] = "Settlement:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = "T+2"
    ws[f'B{row}'].fill = input_fill
    ws[f'B{row}'].border = input_border
    ws[f'C{row}'] = "Settlement convention"
    dv_settle = DataValidation(type="list", formula1='"T+0,T+1,T+2,T+3"', allow_blank=False)
    dv_settle.add(f'B{row}')
    ws.add_data_validation(dv_settle)
    wb.defined_names.add(DefinedName("Sel_Settlement", attr_text=f"Controls!$B${row}"))
    row += 2
    
    # Risk Parameters Section
    ws[f'A{row}'] = "RISK PARAMETERS"
    ws[f'A{row}'].font = section_font_white
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:C{row}')
    row += 1
    
    # KRD Bump Size
    ws[f'A{row}'] = "KRD Bump (bps):"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = 1
    ws[f'B{row}'].fill = input_fill
    ws[f'B{row}'].border = input_border
    ws[f'C{row}'] = "Key rate duration bump"
    dv_krd = DataValidation(type="list", formula1='"0.5,1,5,10"', allow_blank=False)
    dv_krd.add(f'B{row}')
    ws.add_data_validation(dv_krd)
    wb.defined_names.add(DefinedName("Sel_KRDBumpBps", attr_text=f"Controls!$B${row}"))
    row += 1
    
    # Spread for analysis
    ws[f'A{row}'] = "Spread (bps):"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = 0
    ws[f'B{row}'].fill = input_fill
    ws[f'B{row}'].border = input_border
    ws[f'C{row}'] = "Spread for PV calculations"
    wb.defined_names.add(DefinedName("Sel_SpreadBps", attr_text=f"Controls!$B${row}"))
    row += 2
    
    # Bond Information Display (Read-only)
    ws[f'A{row}'] = "BOND INFORMATION"
    ws[f'A{row}'].font = section_font_white
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:C{row}')
    row += 1
    
    info_items = [
        ("Security Name", bond_data['reference']['Security Name']),
        ("Coupon Rate (%)", float(bond_data['reference']['Coupon Rate'])),
        ("Maturity Date", bond_data['schedule']['Maturity Date'].strftime('%d/%m/%Y') 
         if hasattr(bond_data['schedule']['Maturity Date'], 'strftime') 
         else bond_data['schedule']['Maturity Date']),
        ("Notional", 100),
        ("Callable", "Yes" if bond_data.get('call_schedule') else "No"),
    ]
    
    for label, value in info_items:
        ws[f'A{row}'] = label + ":"
        ws[f'B{row}'] = value
        ws[f'B{row}'].fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
        ws[f'B{row}'].border = input_border
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 30
    
    # Freeze panes to keep headers visible
    ws.freeze_panes = 'A4'
    
    # Add calculation status indicator
    row += 2
    ws[f'A{row}'] = "Calc Status:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = "=NOW()"
    ws[f'B{row}'].number_format = "DD/MM/YYYY HH:MM:SS"
    ws[f'C{row}'] = "Last recalculation time"
    wb.defined_names.add(DefinedName("Calc_Status", attr_text=f"Controls!$B${row}"))