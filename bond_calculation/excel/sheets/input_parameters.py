# sheets/input_parameters.py
# Purpose: Build the Input_Parameters sheet

from __future__ import annotations

from datetime import datetime
from typing import Dict

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.workbook.defined_name import DefinedName

from ..styles import header_font, header_fill, input_fill, highlight_fill


def add_input_parameters_sheet(wb: Workbook, bond_data: Dict, price: float, valuation_date: datetime, accrued_interest: float = None) -> None:
    ws = wb.create_sheet("Input_Parameters")
    ws.append(["INPUT PARAMETERS (Modify Blue Cells)"])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])

    ws.append(["Bond Information", "Value", "Description"])
    ws['A3'].font = header_font
    ws['A3'].fill = header_fill

    row = 4
    # Track the worksheet row index for each parameter we add so formulas can
    # reference the correct cells without relying on hard-coded positions
    param_row_by_name: dict[str, int] = {}
    params = [
        ("ISIN", bond_data['reference']['ISIN'], "Bond identifier"),
        ("Security Name", bond_data['reference']['Security Name'], "Bond name"),
        ("Valuation Date", valuation_date, "Analysis date"),
        ("Clean Price", price, "Market price (modifiable)"),
        ("Coupon Rate (%)", float(bond_data['reference']['Coupon Rate']), "Annual coupon rate"),
        ("Maturity Date", bond_data['schedule']['Maturity Date'], "Final payment date"),
        ("Day Count", bond_data['schedule']['Day Basis'], "Day count convention"),
        ("Coupon Frequency", int(bond_data['schedule']['Coupon Frequency']), "Payments per year"),
        ("Notional", 100, "Face value"),
        ("Compounding", "Semiannual", "Yield compounding"),
        ("Callable", "Yes" if bond_data.get('call_schedule') else "No", "Has embedded options"),
        ("Number of Calls", len(bond_data.get('call_schedule') or []), "Call dates available"),
    ]

    for param, value, desc in params:
        if param in ["Valuation Date", "Maturity Date"] and not isinstance(value, (int, float)):
            display_value = value.strftime('%d/%m/%Y') if isinstance(value, datetime) else value
            ws.append([param, display_value, desc])
        else:
            ws.append([param, value, desc])
        # Record the row for this parameter so we can reference it later
        param_row_by_name[param] = row
        if param in ["Clean Price", "Coupon Rate (%)", "Notional"]:
            ws.cell(row=row, column=2).fill = input_fill
        elif param == "Callable" and value == "Yes":
            ws.cell(row=row, column=2).fill = highlight_fill
        row += 1

    # Accrued and Dirty (derived) and named ranges for downstream sheets
    accrued_row = row
    
    # Use accrued from file if available
    if accrued_interest is not None:
        ws.append(["Accrued Interest (File)", accrued_interest, "From sec_accrued.csv"])
        ws.cell(row=accrued_row, column=2).number_format = '0.0000'
    else:
        ws.append(["Accrued Interest (File)", 0.0, "Not found in sec_accrued.csv"])
        ws.cell(row=accrued_row, column=2).number_format = '0.0000'
    
    # Show Excel formula for reference but don't use it
    row += 1
    excel_accrued_row = row
    ws.append(["Accrued (Excel Formula)", None, "Excel calculation (reference only)"])
    # Build Excel formula using COUPPCD/ACCRINT with basis from Assumptions (named range Assump_Basis_Code)
    val_row = param_row_by_name.get("Valuation Date")
    mat_row = param_row_by_name.get("Maturity Date")
    rate_row = param_row_by_name.get("Coupon Rate (%)")
    notional_row = param_row_by_name.get("Notional")
    freq_row = param_row_by_name.get("Coupon Frequency")
    if all(r is not None for r in [val_row, mat_row, rate_row, notional_row, freq_row]):
        ws.cell(row=excel_accrued_row, column=2).value = (
            f"=ACCRINT(COUPPCD(B{val_row},B{mat_row},B{freq_row},Assump_Basis_Code),"
            f"COUPNCD(B{val_row},B{mat_row},B{freq_row},Assump_Basis_Code),"
            f"B{val_row},B{rate_row}/100,B{notional_row},B{freq_row},Assump_Basis_Code,TRUE)"
        )
    else:
        ws.cell(row=excel_accrued_row, column=2).value = 0.0
    ws.cell(row=excel_accrued_row, column=2).number_format = '0.0000'
    from openpyxl.styles import PatternFill
    ws.cell(row=excel_accrued_row, column=2).fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")  # Gray to show it's not used
    
    row += 1
    dirty_row = row
    # Reference the actual Clean Price cell, not the ISIN cell
    clean_price_row = param_row_by_name.get("Clean Price", 7)
    ws.append(["Dirty Price", f"=B{clean_price_row}+B{accrued_row}", "Clean + Accrued (target PV)"])
    ws.cell(row=dirty_row, column=2).number_format = '0.0000'

    # Named ranges: Price_Clean, Price_Accrued, Price_Dirty, Coupon_Rate, Notional, Frequency
    def _add_defined(name: str, ref: str) -> None:
        try:
            dn = DefinedName(name=name, attr_text=ref)
            if hasattr(wb.defined_names, "add"):
                wb.defined_names.add(dn)
            elif hasattr(wb.defined_names, "append"):
                wb.defined_names.append(dn)
        except Exception:
            pass

    _add_defined('Price_Clean', f"Input_Parameters!$B${clean_price_row}")
    _add_defined('Price_Accrued', f"Input_Parameters!$B${accrued_row}")
    _add_defined('Price_Dirty', f"Input_Parameters!$B${dirty_row}")
    
    # Use the actual row numbers from param_row_by_name
    coupon_rate_row = param_row_by_name.get("Coupon Rate (%)", 8)
    notional_row_named = param_row_by_name.get("Notional", 12)
    frequency_row = param_row_by_name.get("Coupon Frequency", 11)
    maturity_date_row = param_row_by_name.get("Maturity Date", 9)
    
    _add_defined('Coupon_Rate', f"Input_Parameters!$B${coupon_rate_row}")
    _add_defined('Notional', f"Input_Parameters!$B${notional_row_named}")
    _add_defined('Frequency', f"Input_Parameters!$B${frequency_row}")
    _add_defined('Maturity', f"Input_Parameters!$B${maturity_date_row}")


