# sheets/volatility_impact.py
# Purpose: Build the Volatility_Impact sheet (illustrative sensitivity table)

from __future__ import annotations

from typing import Dict
from openpyxl import Workbook
from openpyxl.styles import Font
from ..styles import header_font, header_fill, highlight_fill


def add_volatility_impact_sheet(wb: Workbook, python_results: Dict) -> None:
    ws = wb.create_sheet("Volatility_Impact")
    ws.append(["VOLATILITY IMPACT ON OAS"])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])

    ws.append(["How Volatility Affects OAS:"])
    ws['A3'].font = Font(bold=True, size=12)
    ws.append(["Higher volatility → Higher option value → Lower OAS"])
    ws.append(["Lower volatility → Lower option value → Higher OAS"])
    ws.append([])

    ws.append(["Volatility Sensitivity Analysis"])
    ws['A7'].font = Font(bold=True, size=12, color="0066CC")
    headers = ["Volatility", "Option Value", "OAS", "vs 20% Vol"]
    ws.append(headers)
    for i in range(1, len(headers) + 1):
        cell = ws.cell(row=8, column=i)
        cell.font = header_font
        cell.fill = header_fill

    base_oas = python_results.get('oas_standard') or python_results.get('z_spread', 0.0)
    vol_scenarios = [0.10, 0.15, 0.20, 0.25, 0.30]
    for vol in vol_scenarios:
        vol_factor = vol / 0.20
        option_value_scaled = (python_results.get('z_spread', 0.0) - base_oas) * vol_factor if base_oas else 0
        oas_scenario = python_results.get('z_spread', 0.0) - option_value_scaled
        diff_vs_base = (oas_scenario - base_oas) * 10000 if base_oas else 0
        row = ws.max_row + 1
        ws.append([f"{vol*100:.0f}%", f"{option_value_scaled*10000:.1f} bps", f"{oas_scenario*10000:.1f} bps", f"{diff_vs_base:+.1f} bps"])
        if vol == 0.20:
            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = highlight_fill


