# sheets/ytm.py
# Purpose: Build the YTM_Calculations sheet (Python results, Excel methods, XIRR)

from __future__ import annotations

from datetime import datetime
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font


def add_ytm_sheet(
    wb: Workbook,
    python_results: Dict,
    bond_data: Dict,
    cashflows: List[Dict],
    valuation_date: datetime,
) -> None:
    ws = wb.create_sheet("YTM_Calculations")
    ws.append(["YTM CALCULATION - THREE METHODS"])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])

    ws.append(["Method 1: Python SpreadOMatic Results"])
    ws['A3'].font = Font(bold=True, color="0066CC")
    ws.append(["Calculation", "Result", "Status", "Notes"])
    ws.append(["YTM (Python):", python_results['ytm'], "Calculated" if python_results.get('calculated', False) else "Default", "Using existing SpreadOMatic functions"])
    ws.append(["Z-Spread (bps):", python_results['z_spread'] * 10000, "Calculated" if python_results.get('calculated', False) else "Default", f"{python_results['z_spread'] * 10000:.2f} basis points"])
    ws.append(["G-Spread (bps):", python_results['g_spread'] * 10000, "Calculated" if python_results.get('calculated', False) else "Default", f"{python_results['g_spread'] * 10000:.2f} basis points"])

    # Show OAS results if callable
    if bond_data.get('call_schedule'):
        ws.append(["OAS Standard (bps):", python_results['oas_standard'] * 10000 if python_results['oas_standard'] else "N/A", "Calculated" if python_results['oas_standard'] else "Failed", "Single call, 20% volatility"])
        ws.append(["OAS Enhanced (bps):", python_results['oas_enhanced'] * 10000 if python_results['oas_enhanced'] else "N/A", "Calculated" if python_results['oas_enhanced'] else "N/A", "All calls, calibrated volatility"])

    # Format some number cells
    ws.cell(row=5, column=2).number_format = '0.0000%'
    ws.cell(row=6, column=2).number_format = '0.00'
    ws.cell(row=7, column=2).number_format = '0.00'
    if bond_data.get('call_schedule'):
        ws.cell(row=8, column=2).number_format = '0.00'
        ws.cell(row=9, column=2).number_format = '0.00'

    ws.append([])
    ws.append(["Method 2: Excel YIELD Function"])
    ws['A' + str(ws.max_row)].font = Font(bold=True, color="0066CC")

    if cashflows:
        first_cf = cashflows[0]
        last_cf = cashflows[-1]
        settlement_uk_format = valuation_date.strftime('%d/%m/%Y')
        maturity_uk_format = last_cf['date'].strftime('%d/%m/%Y')
        ws.append(["Settlement Date:", settlement_uk_format])
        settle_row = ws.max_row
        ws.cell(row=settle_row, column=2).number_format = 'dd/mm/yyyy'
        ws.append(["Maturity Date:", maturity_uk_format])
        maturity_row = ws.max_row
        ws.cell(row=maturity_row, column=2).number_format = 'dd/mm/yyyy'

        ws.append([])
        ws.append(["XIRR Calculation Data:"])
        ws.append(["Date", "Cashflow"])
        xirr_header_row = ws.max_row
        xirr_values_start_row = xirr_header_row + 1
        # Initial cashflow uses Dirty Price (negative outflow at settlement)
        ws.append([valuation_date, f"=-Price_Dirty"])
        ws.cell(row=xirr_values_start_row, column=1).number_format = 'dd/mm/yyyy'
        for i, cf in enumerate(cashflows):
            ws.append([cf['date'], cf['total']])
            ws.cell(row=xirr_values_start_row + i + 1, column=1).number_format = 'dd/mm/yyyy'
        last_cf_row = xirr_values_start_row + len(cashflows)
        ws.append([])
        # XIRR over the values (excluding headers)
        ws.append(["Excel XIRR (daily):", f'=XIRR(B{xirr_values_start_row}:B{last_cf_row},A{xirr_values_start_row}:A{last_cf_row})'])
        xirr_daily_row = ws.max_row
        ws.cell(row=xirr_daily_row, column=2).number_format = '0.0000%'
        # Convert the effective annual XIRR to a nominal rate using coupon frequency
        ws.append(["Excel NOMINAL (annual, using frequency):", f"=NOMINAL(B{xirr_daily_row}, Assump_Frequency)"])
        ws.cell(row=ws.max_row, column=2).number_format = '0.0000%'
        # Native Excel YIELD using clean price and assumptions
        ws.append([
            "Excel YIELD (clean, nominal):",
            f"=YIELD(B{settle_row},B{maturity_row},Coupon_Rate/100,Price_Clean,100,Frequency,Assump_Basis_Code)"
        ])
        ws.cell(row=ws.max_row, column=2).number_format = '0.0000%'


