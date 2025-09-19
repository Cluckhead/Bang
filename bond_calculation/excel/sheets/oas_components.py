# sheets/oas_components.py
# Purpose: Build the OAS_Components sheet (summary of inputs/components)

from __future__ import annotations

from datetime import datetime
from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font

from ..styles import header_font, header_fill


def add_oas_components_sheet(
    wb: Workbook,
    bond_data: Dict,
    cashflows: List[Dict],
    curve_data: Tuple[List[float], List[float]],
    price: float,
    valuation_date: datetime,
    python_results: Dict,
) -> None:
    ws = wb.create_sheet("OAS_Components")
    ws.append(["OAS CALCULATION COMPONENTS"])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])

    ws.append(["Key Components for OAS Calculation:"])
    ws['A3'].font = Font(bold=True, size=12)
    ws.append([])

    headers = ["Component", "Value/Status", "Purpose"]
    ws.append(headers)
    for i in range(1, len(headers) + 1):
        cell = ws.cell(row=5, column=i)
        cell.font = header_font
        cell.fill = header_fill

    def _fmt(val):
        try:
            return f"{val:.4f}"
        except Exception:
            return str(val)

    components = [
        ("Clean Price", f"{price:.2f}", "Current market price"),
        ("Yield Curve", f"{len(curve_data[0])} points", "Risk-free discount rates"),
        ("Cash Flows", f"{len(cashflows)} payments", "Bond payment schedule"),
        (
            "Call Schedule",
            f"{len(bond_data.get('call_schedule') or [])} calls" if bond_data.get('call_schedule') else "None",
            "Embedded options",
        ),
        ("YTM", f"{python_results.get('ytm', 0)*100:.3f}%", "Yield to maturity"),
        ("Z-Spread", f"{python_results.get('z_spread', 0)*10000:.1f} bps", "Base spread over curve"),
    ]
    for name, value, purpose in components:
        ws.append([name, value, purpose])


