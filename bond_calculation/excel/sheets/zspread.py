# sheets/zspread.py
# Purpose: Build the ZSpread_Calculations sheet

from __future__ import annotations

from typing import Dict, List
from openpyxl import Workbook
from ..styles import header_font, header_fill, input_fill


def add_zspread_sheet(wb: Workbook, cashflows: List[Dict]) -> None:
    ws = wb.create_sheet("ZSpread_Calculations")
    ws.append(["Z-SPREAD CALCULATION WITH FORMULAS"])
    from openpyxl.styles import Font
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])

    ws.append(["Z-Spread Input (bps):", 50, "← Modify to see impact"])
    ws.cell(row=3, column=2).fill = input_fill

    ws.append([])
    ws.append(["Note: Linear Interpolation uses FORECAST function"])
    ws['A5'].font = Font(italic=True)
    ws.append([])

    # Table headers
    headers = [
        "CF #",
        "Time",
        "Cashflow",
        "Interp Zero Rate",
        "Zero + Spread",
        "Discount Factor",
        "Present Value",
    ]
    ws.append(headers)
    for i in range(1, len(headers) + 1):
        cell = ws.cell(row=7, column=i)
        cell.font = header_font
        cell.fill = header_fill

    # We will reference dynamic ranges on Cashflows and Yield_Curve
    # After headers in row 7, next append will be row 8
    ws.append([])  # Add empty row 8 for spacing
    
    # Now data starts at row 9
    start_row = 9
    num_rows = max(0, len(cashflows))
    for i in range(1, num_rows + 1):
        current_row = start_row + i - 1  # This is the row we're writing to
        # For referencing Cashflows sheet, we need i+1 because their data starts at row 2
        cashflow_ref_row = i + 1
        
        ws.append([
            i,  # CF number
            f"=Cashflows!D{cashflow_ref_row}",  # Time from Cashflows sheet
            f"=Cashflows!I{cashflow_ref_row}",  # Amount from Cashflows sheet
            f"=FORECAST.LINEAR(B{current_row},Curve_Rates,Curve_Terms)",  # Interp using THIS row's time
            f"=D{current_row}+(($B$3+Assump_ZShift_bps)/100)",  # Zero + spread using THIS row's interp rate
            f'=IF(Sel_Compounding="Continuous",EXP(-E{current_row}/100*B{current_row}),1/(1+E{current_row}/100/Sel_Freq)^(Sel_Freq*B{current_row}))',  # DF using THIS row
            f"=C{current_row}*F{current_row}",  # PV using THIS row's CF and DF
        ])

    last_calc_row = start_row + max(0, num_rows) - 1
    ws.append([])
    total_pv_formula = f"=IF({last_calc_row}<{start_row},0,SUM(G{start_row}:G{last_calc_row}))"
    ws.append(["", "", "Total PV:", "", "", "", total_pv_formula]) 
    ws.append(["", "", "Target Price:", "", "", "", f"=Price_Dirty"]) 
    # Error should subtract Target Price from Total PV; reference prior two rows to avoid self-reference
    # Current row will be two rows below Total PV, one below Target Price
    ws.append(["", "", "Error:", "", "", "", f"=G{ws.max_row-1}-G{ws.max_row}"]) 


