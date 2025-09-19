# sheets/summary_comparison.py
# Purpose: Build enhanced Summary_Comparison sheet with Python vs Excel comparison

from __future__ import annotations

from typing import Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.workbook.defined_name import DefinedName

from ..styles import header_font, header_fill


def add_summary_comparison_sheet(wb: Workbook, bond_data: Dict, python_results: Dict) -> None:
    """
    Create an enhanced Summary_Comparison sheet with Python vs Excel comparison,
    including differences, percentage differences, and status indicators.
    """
    ws = wb.create_sheet("Summary_Comparison")
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "SPREADOMATIC vs EXCEL COMPARISON"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    ws.append([])
    headers = ["Metric", "SpreadOMatic (Python)", "Excel", "Difference", "% Diff", "Tolerance", "Status", "Notes"]
    ws.append(headers)
    for i in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=i)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Define tolerances
    tolerances = {
        'YTM': 0.0001,  # 1 bp
        'Z-Spread': 0.1,  # 0.1 bp
        'G-Spread': 0.1,  # 0.1 bp
        'OAS': 0.5,  # 0.5 bp
        'Duration': 0.001,  # 0.001 years
        'Convexity': 0.01,  # 0.01
        'PV': 0.01,  # $0.01
    }
    
    row = 4
    
    # YTM Section
    ws[f'A{row}'] = "YTM (%)"
    ws[f'B{row}'] = f"{python_results['ytm']:.6f}"
    ws[f'C{row}'] = "=YTM_Calculations!B5"  # Reference to Excel YTM calculation
    ws[f'D{row}'] = f"=C{row}-B{row}"
    ws[f'E{row}'] = f"=IF(B{row}<>0,D{row}/B{row}*100,0)"
    ws[f'F{row}'] = tolerances['YTM']
    ws[f'G{row}'] = f'=IF(ABS(D{row})<=F{row},"✓","✗")'
    ws[f'H{row}'] = "Newton-Raphson vs Excel"
    row += 1
    
    # Z-Spread
    ws[f'A{row}'] = "Z-Spread (bps)"
    ws[f'B{row}'] = f"{python_results['z_spread']*10000:.2f}"
    ws[f'C{row}'] = "=ZSpread_Calculations!B3"  # Reference to Excel Z-spread input/calc
    ws[f'D{row}'] = f"=C{row}-B{row}"
    ws[f'E{row}'] = f"=IF(B{row}<>0,D{row}/B{row}*100,0)"
    ws[f'F{row}'] = tolerances['Z-Spread']
    ws[f'G{row}'] = f'=IF(ABS(D{row})<=F{row},"✓","✗")'
    ws[f'H{row}'] = "Goal Seek required"
    row += 1
    
    # G-Spread
    if python_results.get('g_spread') is not None:
        ws[f'A{row}'] = "G-Spread (bps)"
        ws[f'B{row}'] = f"{python_results['g_spread']*10000:.2f}"
        ws[f'C{row}'] = "=(YTM_Calculations!B5-Yield_Curve!B10)*100"  # YTM minus govt rate
        ws[f'D{row}'] = f"=C{row}-B{row}"
        ws[f'E{row}'] = f"=IF(B{row}<>0,D{row}/B{row}*100,0)"
        ws[f'F{row}'] = tolerances['G-Spread']
        ws[f'G{row}'] = f'=IF(ABS(D{row})<=F{row},"✓","✗")'
        ws[f'H{row}'] = "YTM vs govt curve"
        row += 1
    
    # Duration Metrics
    ws[f'A{row}'] = "Effective Duration"
    ws[f'B{row}'] = f"{python_results.get('effective_duration', 0):.6f}"
    ws[f'C{row}'] = "=Effective_Duration!B5"  # Reference to Excel calculation
    ws[f'D{row}'] = f"=C{row}-B{row}"
    ws[f'E{row}'] = f"=IF(B{row}<>0,D{row}/B{row}*100,0)"
    ws[f'F{row}'] = tolerances['Duration']
    ws[f'G{row}'] = f'=IF(ABS(D{row})<=F{row},"✓","✗")'
    ws[f'H{row}'] = "±10bp parallel shift"
    row += 1
    
    ws[f'A{row}'] = "Modified Duration"
    ws[f'B{row}'] = f"{python_results.get('modified_duration', 0):.6f}"
    ws[f'C{row}'] = "=Duration_Summary!B6"  # Reference to Excel calculation
    ws[f'D{row}'] = f"=C{row}-B{row}"
    ws[f'E{row}'] = f"=IF(B{row}<>0,D{row}/B{row}*100,0)"
    ws[f'F{row}'] = tolerances['Duration']
    ws[f'G{row}'] = f'=IF(ABS(D{row})<=F{row},"✓","✗")'
    ws[f'H{row}'] = "Analytical formula"
    row += 1
    
    # Convexity
    ws[f'A{row}'] = "Convexity"
    ws[f'B{row}'] = f"{python_results.get('convexity', 0):.4f}"
    ws[f'C{row}'] = "=Convexity!B5"  # Reference to Excel calculation
    ws[f'D{row}'] = f"=C{row}-B{row}"
    ws[f'E{row}'] = f"=IF(B{row}<>0,D{row}/B{row}*100,0)"
    ws[f'F{row}'] = tolerances['Convexity']
    ws[f'G{row}'] = f'=IF(ABS(D{row})<=F{row},"✓","✗")'
    ws[f'H{row}'] = "Second derivative"
    row += 1
    
    # OAS (if callable)
    if bond_data.get('call_schedule'):
        ws[f'A{row}'] = "OAS Standard (bps)"
        oas_val = python_results.get('oas_standard')
        if oas_val and oas_val != "N/A":
            ws[f'B{row}'] = f"{oas_val*10000:.2f}"
        else:
            ws[f'B{row}'] = "N/A"
        ws[f'C{row}'] = "=OAS_Calculation!B5"  # Reference to Excel OAS
        ws[f'D{row}'] = f'=IF(AND(ISNUMBER(B{row}),ISNUMBER(C{row})),C{row}-B{row},"N/A")'
        ws[f'E{row}'] = f'=IF(AND(ISNUMBER(B{row}),ISNUMBER(C{row}),B{row}<>0),D{row}/B{row}*100,"N/A")'
        ws[f'F{row}'] = tolerances['OAS']
        ws[f'G{row}'] = f'=IF(ISNUMBER(D{row}),IF(ABS(D{row})<=F{row},"✓","✗"),"N/A")'
        ws[f'H{row}'] = "Black Model"
        row += 1
        
        ws[f'A{row}'] = "OAS Enhanced (bps)"
        oas_enh = python_results.get('oas_enhanced')
        if oas_enh and oas_enh != "N/A":
            ws[f'B{row}'] = f"{oas_enh*10000:.2f}"
        else:
            ws[f'B{row}'] = "N/A"
        ws[f'C{row}'] = "=OAS_Components!B10"  # Reference to Excel OAS Enhanced
        ws[f'D{row}'] = f'=IF(AND(ISNUMBER(B{row}),ISNUMBER(C{row})),C{row}-B{row},"N/A")'
        ws[f'E{row}'] = f'=IF(AND(ISNUMBER(B{row}),ISNUMBER(C{row}),B{row}<>0),D{row}/B{row}*100,"N/A")'
        ws[f'F{row}'] = tolerances['OAS']
        ws[f'G{row}'] = f'=IF(ISNUMBER(D{row}),IF(ABS(D{row})<=F{row},"✓","✗"),"N/A")'
        ws[f'H{row}'] = "Hull-White MC"
        row += 1
    
    # Present Value
    ws[f'A{row}'] = "Present Value"
    ws[f'B{row}'] = f"{python_results.get('present_value', 0):.4f}"
    ws[f'C{row}'] = "=ZSpread_Calculations!G20"  # Total PV from Z-spread sheet
    ws[f'D{row}'] = f"=C{row}-B{row}"
    ws[f'E{row}'] = f"=IF(B{row}<>0,D{row}/B{row}*100,0)"
    ws[f'F{row}'] = tolerances['PV']
    ws[f'G{row}'] = f'=IF(ABS(D{row})<=F{row},"✓","✗")'
    ws[f'H{row}'] = "Discounted cashflows"
    row += 2
    
    # Summary Statistics
    ws[f'A{row}'] = "SUMMARY STATISTICS"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    
    # Count metrics
    metrics_range = f"G4:G{row-3}"
    ws[f'A{row}'] = "Metrics Passed:"
    ws[f'B{row}'] = f'=COUNTIF({metrics_range},"✓")'
    ws[f'B{row}'].font = Font(bold=True, color="27AE60")
    row += 1
    
    ws[f'A{row}'] = "Metrics Failed:"
    ws[f'B{row}'] = f'=COUNTIF({metrics_range},"✗")'
    ws[f'B{row}'].font = Font(bold=True, color="E74C3C")
    row += 1
    
    ws[f'A{row}'] = "Total Metrics:"
    ws[f'B{row}'] = f'=COUNTA({metrics_range})-COUNTIF({metrics_range},"N/A")'
    ws[f'B{row}'].font = Font(bold=True)
    row += 1
    
    ws[f'A{row}'] = "Pass Rate:"
    ws[f'B{row}'] = f'=IF(B{row-1}>0,B{row-3}/B{row-1}*100,0)'
    ws[f'B{row}'].number_format = '0.0"%"'
    ws[f'B{row}'].font = Font(bold=True)
    row += 2
    
    # Validation Flags
    ws[f'A{row}'] = "VALIDATION FLAGS"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    
    flags = [
        ("Curve Extrapolation", '=IF(MAX(Cashflows!D:D)>MAX(Curve_Terms),"WARNING: Extrapolating beyond curve","OK")', 
         "Check if cashflows extend beyond curve"),
        ("Negative Rates", '=IF(COUNTIF(Curve_Rates,"<0")>0,"WARNING: Negative rates present","OK")', 
         "Check for negative yields"),
        ("Data Freshness", f'=IF(Controls!B8="{python_results.get("valuation_date", "")}","Data synchronized","WARNING: Date mismatch")', 
         "Verify valuation dates match"),
        ("Solver Convergence", '=IF(ABS(ZSpread_Calculations!G22)<0.01,"Converged","WARNING: Check Goal Seek")', 
         "Z-spread convergence check"),
    ]
    
    for flag_name, formula, description in flags:
        ws[f'A{row}'] = flag_name
        ws[f'B{row}'] = formula
        ws[f'C{row}'] = description
        
        # Apply conditional formatting
        if "WARNING" in formula:
            ws[f'B{row}'].font = Font(color="E74C3C")
        row += 1
    
    # Apply conditional formatting to Status column
    for status_row in range(4, row-10):  # Adjust range as needed
        pass_rule = CellIsRule(operator='equal', formula=['"✓"'], 
                               font=Font(color="27AE60", bold=True))
        fail_rule = CellIsRule(operator='equal', formula=['"✗"'], 
                               font=Font(color="E74C3C", bold=True))
        ws.conditional_formatting.add(f'G{status_row}', pass_rule)
        ws.conditional_formatting.add(f'G{status_row}', fail_rule)
    
    # Format difference columns
    for diff_row in range(4, row-10):
        ws[f'D{diff_row}'].number_format = '0.000000'
        ws[f'E{diff_row}'].number_format = '0.00"%"'
    
    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 25
    
    # Freeze panes
    ws.freeze_panes = 'A4'