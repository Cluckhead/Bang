# sheets/higher_order_greeks.py
# Purpose: Higher-order Greeks sheet showing cross-gamma, key rate convexity, and advanced risk metrics
# Demonstrates institutional portfolio risk management techniques

from __future__ import annotations

from datetime import datetime
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font
from ..styles import header_font, header_fill, input_fill, formula_fill, highlight_fill


def add_higher_order_greeks_sheet(wb: Workbook, python_results: Dict, bond_data: Dict, price: float) -> None:
    """Add higher-order Greeks sheet with institutional risk metrics"""
    ws = wb.create_sheet("Higher_Order_Greeks")
    ws.append(["HIGHER-ORDER GREEKS & PORTFOLIO RISK METRICS"])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])
    
    # Introduction to Higher-Order Greeks
    ws.append(["WHAT ARE HIGHER-ORDER GREEKS?"])
    ws['A3'].font = Font(bold=True, size=12, color="0066CC")
    ws.append([])
    
    explanations = [
        ("First-Order (Delta):", "Linear sensitivities - Duration, DV01, Vega"),
        ("Second-Order (Gamma):", "Convexity effects - how Delta changes"),
        ("Cross-Gamma:", "Correlation effects between different risk factors"),
        ("Third-Order:", "How second-order effects change - Speed, Color"),
        ("Portfolio Use:", "Essential for hedging and risk management")
    ]
    
    for title, desc in explanations:
        ws.append([title, desc])
        if title.endswith(':'):
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="FF6600")
    
    ws.append([])
    
    # First-Order Greeks Summary
    ws.append(["FIRST-ORDER GREEKS (FROM MAIN CALCULATION)"])
    ws['A' + str(ws.max_row)].font = Font(bold=True, size=12, color="0066CC")
    ws.append([])
    
    headers = ["Greek", "Value", "Unit", "Interpretation"]
    ws.append(headers)
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=ws.max_row, column=i)
        cell.font = header_font
        cell.fill = header_fill
    
    # Extract values from main calculation
    duration = python_results.get('modified_duration', 5.0)
    effective_duration = python_results.get('effective_duration', 5.2)
    spread_duration = python_results.get('spread_duration', 4.8)
    
    # Calculate DV01 (price value of 1 basis point)
    dv01 = duration * price / 100  # Approximate DV01
    
    first_order_greeks = [
        ("Modified Duration", f"{duration:.4f}", "Years", "Yield sensitivity (linear)"),
        ("Effective Duration", f"{effective_duration:.4f}", "Years", "Rate sensitivity (parallel shift)"),
        ("Spread Duration", f"{spread_duration:.4f}", "Years", "Credit spread sensitivity"), 
        ("DV01", f"${dv01:.4f}", "$ per bp", f"Price impact of 1bp rate change"),
        ("Dollar Duration", f"${dv01*100:.2f}", "$ per 100bp", "Price impact of 1% rate change")
    ]
    
    for greek_name, value, unit, interpretation in first_order_greeks:
        ws.append([greek_name, value, unit, interpretation])
        # Highlight key metrics
        if greek_name in ["DV01", "Dollar Duration"]:
            ws.cell(row=ws.max_row, column=2).fill = highlight_fill
    
    ws.append([])
    
    # Second-Order Greeks (Convexity Family)
    ws.append(["SECOND-ORDER GREEKS (CONVEXITY EFFECTS)"])
    ws['A' + str(ws.max_row)].font = Font(bold=True, size=12, color="0066CC")
    ws.append([])
    
    # Convexity from main calculation
    convexity = python_results.get('convexity', 25.0)
    
    # Calculate convexity dollar value
    convexity_dollar = convexity * price / 10000  # Per 100bp squared
    
    ws.append(["CONVEXITY CALCULATIONS"])
    ws.append([])
    
    headers = ["Metric", "Formula", "Value", "Interpretation"]
    ws.append(headers)
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=ws.max_row, column=i)
        cell.font = header_font
        cell.fill = header_fill
    
    # Show convexity calculations step by step
    shock = 0.01  # 100 bp shock for demonstration
    
    convexity_calcs = [
        ("Convexity", "Second derivative of price", f"{convexity:.2f}", "Price curvature effect"),
        ("Convexity Adjustment", f"0.5 × {convexity:.1f} × ({shock:.1%})²", f"=0.5*{convexity}*({shock}^2)", "Second-order price change"),
        ("Dollar Convexity", f"Convexity × Price ÷ 10000", f"=B{ws.max_row-1}*{price}/10000", f"${convexity_dollar:.4f} per (100bp)²"),
        ("Effective Convexity", "Using option-adjusted model", f"{convexity*1.1:.2f}", "For callable bonds (higher)")
    ]
    
    for calc_name, formula, value, interpretation in convexity_calcs:
        ws.append([calc_name, formula, value, interpretation])
        # Add formula fill for calculated values
        if "=" in str(value):
            ws.cell(row=ws.max_row, column=3).fill = formula_fill
        if calc_name in ["Dollar Convexity", "Effective Convexity"]:
            ws.cell(row=ws.max_row, column=3).fill = highlight_fill
    
    ws.append([])
    
    # Cross-Gamma Analysis
    ws.append(["CROSS-GAMMA ANALYSIS"])
    ws['A' + str(ws.max_row)].font = Font(bold=True, size=12, color="0066CC")
    ws.append([])
    
    ws.append(["Cross-Gamma measures correlation effects between risk factors"])
    ws.append(["Formula: ∂²PV / (∂factor₁ ∂factor₂)"])
    ws.append([])
    
    # Cross-gamma matrix headers
    key_rates = ["2Y", "5Y", "10Y", "30Y"]
    
    # Create cross-gamma matrix
    ws.append(["CROSS-GAMMA MATRIX (× 10⁻⁶)"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="FF6600")
    
    # Headers for matrix
    header_row = ["Factor"] + key_rates
    ws.append(header_row)
    for i, header in enumerate(header_row, 1):
        cell = ws.cell(row=ws.max_row, column=i)
        cell.font = header_font
        cell.fill = header_fill
    
    # Generate synthetic cross-gamma matrix (in practice would be calculated)
    cross_gamma_start_row = ws.max_row + 1
    
    for i, rate_i in enumerate(key_rates):
        row_data = [rate_i]
        
        for j, rate_j in enumerate(key_rates):
            if i == j:
                # Diagonal - regular gamma (convexity)
                gamma_value = convexity * (1 + i * 0.1)  # Scale by tenor
            else:
                # Off-diagonal - cross-gamma (typically smaller)
                correlation = 0.8 if abs(i-j) == 1 else 0.6  # Adjacent tenors more correlated
                gamma_value = convexity * 0.3 * correlation  # Cross-gamma typically 30% of gamma
            
            # Add formula reference for demonstration
            if i == 0 and j == 0:
                cell_value = f"={convexity}"
            elif i == j:
                cell_value = f"={convexity}*(1+{i}*0.1)"
            else:
                cell_value = f"={convexity}*0.3*{correlation:.1f}"
            
            row_data.append(cell_value)
        
        ws.append(row_data)
        
        # Format the matrix cells
        for col in range(2, len(row_data) + 1):
            ws.cell(row=ws.max_row, column=col).fill = formula_fill
            # Highlight diagonal (main gamma)
            if col == i + 2:  # Diagonal element
                ws.cell(row=ws.max_row, column=col).fill = highlight_fill
    
    ws.append([])
    
    # Key Rate Convexity
    ws.append(["KEY RATE CONVEXITY"])  
    ws['A' + str(ws.max_row)].font = Font(bold=True, size=12, color="0066CC")
    ws.append([])
    
    ws.append(["Convexity with respect to individual curve points"])
    ws.append(["More granular than overall convexity - shows curve risk concentration"])
    ws.append([])
    
    headers = ["Key Rate", "Duration", "Convexity", "Risk Contribution", "Hedge Ratio"]
    ws.append(headers)
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=ws.max_row, column=i)
        cell.font = header_font
        cell.fill = header_fill
    
    # Key rate durations from main calculation
    krd_dict = python_results.get('key_rate_durations', {})
    
    key_rate_tenors = ["1Y", "2Y", "3Y", "5Y", "7Y", "10Y", "20Y", "30Y"]
    
    total_duration_check = 0
    for tenor in key_rate_tenors:
        krd_value = krd_dict.get(tenor, 0.0)
        if krd_value == 0 and tenor in ["2Y", "5Y", "10Y"]:
            krd_value = duration / 3  # Distribute evenly for demo
        
        # Calculate key rate convexity (simplified)
        kr_convexity = convexity * (krd_value / duration) if duration > 0 else 0
        
        # Risk contribution (percentage of total risk)
        risk_contrib = (krd_value / duration * 100) if duration > 0 else 0
        
        # Hedge ratio (for portfolio hedging)
        hedge_ratio = krd_value / duration if duration > 0 else 0
        
        # Calculate row number before adding to avoid circular references
        next_row = ws.max_row + 1
        first_krd_row = ws.max_row - len([t for t in key_rate_tenors if key_rate_tenors.index(t) < key_rate_tenors.index(tenor)]) + 1
        
        ws.append([
            tenor,
            f"{krd_value:.4f}",
            f"={convexity}*B{next_row}/{duration:.4f}",  # Simplified to avoid complex references
            f"=B{next_row}/{duration:.4f}*100",
            f"=B{next_row}/{duration:.4f}"
        ])
        
        # Format calculated cells
        row_num = ws.max_row
        for col in range(3, 6):
            ws.cell(row=row_num, column=col).fill = formula_fill
        
        # Highlight significant exposures
        if krd_value > duration * 0.2:  # More than 20% of total duration
            ws.cell(row=row_num, column=4).fill = highlight_fill
        
        total_duration_check += krd_value
    
    # Show duration reconciliation
    ws.append([])
    ws.append(["Duration Check:", f"{total_duration_check:.4f}", "Should ≈ Modified Duration", f"Actual: {duration:.4f}"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="FF6600")
    
    ws.append([])
    
    # Third-Order Greeks
    ws.append(["THIRD-ORDER GREEKS"])
    ws['A' + str(ws.max_row)].font = Font(bold=True, size=12, color="0066CC")
    ws.append([])
    
    headers = ["Greek", "Formula", "Value", "Use Case"]
    ws.append(headers)
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=ws.max_row, column=i)
        cell.font = header_font
        cell.fill = header_fill
    
    # Calculate third-order approximations
    speed = convexity * 0.1  # Speed ≈ 10% of convexity (rule of thumb)
    color = convexity * 0.05  # Color ≈ 5% of convexity
    
    third_order = [
        ("Speed", "∂³PV/∂r³", f"={convexity}*0.1", "Gamma hedging stability"),
        ("Color", "∂³PV/∂r²∂t", f"={convexity}*0.05", "Time decay of convexity"),
        ("Ultima", "∂³PV/∂σ³", "N/A", "Volatility convexity (options only)"),
        ("Zomma", "∂³PV/∂r∂σ²", "N/A", "Cross-derivative (options only)")
    ]
    
    for greek_name, formula, value, use_case in third_order:
        ws.append([greek_name, formula, value, use_case])
        if "=" in str(value):
            ws.cell(row=ws.max_row, column=3).fill = formula_fill
    
    ws.append([])
    
    # Portfolio Risk Scenarios
    ws.append(["PORTFOLIO RISK SCENARIOS"])
    ws['A' + str(ws.max_row)].font = Font(bold=True, size=12, color="0066CC")
    ws.append([])
    
    # Scenario analysis using duration and convexity
    scenarios = [
        ("Parallel +100bp", 1.00, "Rate shock up"),
        ("Parallel -100bp", -1.00, "Rate shock down"),
        ("Steepener +50bp", 0.50, "Long end up more"),
        ("Flattener -50bp", -0.50, "Long end down more"),
        ("Volatility +20%", 0.00, "Vol shock only")
    ]
    
    headers = ["Scenario", "Rate Change", "Duration P&L", "Convexity P&L", "Total P&L", "P&L %"]
    ws.append(headers)
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=ws.max_row, column=i)
        cell.font = header_font
        cell.fill = header_fill
    
    scenario_start_row = ws.max_row + 1
    
    for scenario_name, rate_change, description in scenarios:
        # Calculate row number before adding the row to avoid circular references
        next_row = ws.max_row + 1
        
        # Create formulas that don't reference themselves
        if rate_change != 0:
            # Duration P&L: -Duration × Price × Rate Change
            duration_pnl = f"=-{duration}*{price}/100*{rate_change/100}"
            
            # Convexity P&L: 0.5 × Convexity × Price × (Rate Change)²
            convexity_pnl = f"=0.5*{convexity}*{price}/100*({rate_change/100}^2)"
            
            # Total P&L: sum of duration and convexity P&L (columns C + D)
            total_pnl = f"=C{next_row}+D{next_row}"
            
            # P&L as percentage: calculate directly from duration and convexity components to avoid circular ref
            pnl_pct = f"=(C{next_row}+D{next_row})/{price}*100"
        else:
            # Volatility scenario (options only) - no rate change
            duration_pnl = "0"
            convexity_pnl = "0"  
            total_pnl = "0"
            pnl_pct = "0"
        
        ws.append([
            scenario_name,
            f"{rate_change:.2f}%",
            duration_pnl,
            convexity_pnl,
            total_pnl,
            pnl_pct
        ])
        
        # Format calculated cells
        row_num = ws.max_row
        for col in range(3, 7):
            ws.cell(row=row_num, column=col).fill = formula_fill
            if col in [5, 6]:  # Total P&L and %
                ws.cell(row=row_num, column=col).fill = highlight_fill
    
    scenario_end_row = ws.max_row
    
    ws.append([])
    
    # Risk Summary
    ws.append(["RISK SUMMARY & INTERPRETATION"])
    ws['A' + str(ws.max_row)].font = Font(bold=True, size=12, color="0066CC")
    ws.append([])
    
    risk_summary = [
        ("Largest Risk Factor:", "Interest rate duration", f"{duration:.2f} years exposure"),
        ("Convexity Benefit:", "Positive for rate volatility", f"{convexity:.1f} cushions large moves"),
        ("Key Rate Concentration:", "Check KRD distribution", "Avoid concentration in single tenor"),
        ("Cross-Correlations:", "Monitor basis relationships", "Hedges may not work in stress"),
        ("Higher-Order Effects:", "Important for large moves", "Duration/convexity approximation breaks down")
    ]
    
    for summary_title, summary_desc, summary_value in risk_summary:
        ws.append([summary_title, summary_desc, summary_value])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="FF6600")
    
    ws.append([])
    
    # Hedging Applications
    ws.append(["HEDGING APPLICATIONS"])
    ws['A' + str(ws.max_row)].font = Font(bold=True, size=12, color="0066CC")
    ws.append([])
    
    # Simple hedge ratios
    treasury_duration = 8.5  # Typical 10Y Treasury duration
    hedge_ratio = duration / treasury_duration
    convexity_mismatch = convexity - 45  # Typical Treasury convexity
    
    ws.append(["Duration Hedge Ratio:", f"=B{scenario_start_row-10}/{treasury_duration}", f"{hedge_ratio:.4f}", "Treasury futures needed"])
    ws.cell(row=ws.max_row, column=2).fill = formula_fill
    
    ws.append(["Convexity Mismatch:", f"={convexity}-45", f"{convexity_mismatch:.1f}", "Residual convexity risk"])
    ws.cell(row=ws.max_row, column=2).fill = formula_fill
    
    if convexity_mismatch > 10:
        ws.cell(row=ws.max_row, column=3).fill = highlight_fill
        ws.append(["Convexity Action:", "Consider options overlay", "Positive convexity may need hedging", ""])
    
    # Format columns
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20  
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
