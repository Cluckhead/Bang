# sheets/duration_summary.py
# Purpose: Build the Duration_Summary sheet

from __future__ import annotations

from typing import Dict
from openpyxl import Workbook
from openpyxl.styles import Font
from ..styles import header_font, header_fill


def add_duration_summary_sheet(wb: Workbook, python_results: Dict) -> None:
    ws = wb.create_sheet("Duration_Summary")
    ws.append(["DURATION METRICS SUMMARY"])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])

    headers = ["Metric", "Python Value", "Excel Formula", "Description"]
    ws.append(headers)
    for i in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=i)
        cell.font = header_font
        cell.fill = header_fill

    metrics = [
        ("Effective Duration", python_results.get('effective_duration', 0.0), "=Effective_Duration!B13", "Price sensitivity to parallel rate shifts"),
        ("Modified Duration", python_results.get('modified_duration', 0.0), "=Effective_Duration!B14", "Effective duration adjusted for yield"),
        ("Convexity", python_results.get('convexity', 0.0), "=Convexity!B3", "Second-order price sensitivity"),
        ("Spread Duration", python_results.get('spread_duration', 0.0), "See Python calc", "Sensitivity to spread changes"),
    ]

    row = 4
    for name, val, ref, desc in metrics:
        ws.append([name, val, ref, desc])
        ws.cell(row=row, column=2).number_format = '0.0000'
        if name == "Convexity":
            ws.cell(row=row, column=2).number_format = '0.00'
        row += 1


