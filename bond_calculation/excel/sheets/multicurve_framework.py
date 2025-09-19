# sheets/multicurve_framework.py
# Purpose: Multi-curve framework sheet showing OIS discounting vs LIBOR/SOFR projection
# Demonstrates post-2008 institutional approach to yield curve construction

from __future__ import annotations

from datetime import datetime
from typing import Dict, Tuple, List

from openpyxl import Workbook
from openpyxl.styles import Font
from ..styles import header_font, header_fill, input_fill, formula_fill, highlight_fill


def add_multicurve_framework_sheet(wb: Workbook, curve_data: Tuple[List[float], List[float]], 
                                 bond_data: Dict, python_results: Dict) -> None:
    """Add multi-curve framework sheet showing institutional curve separation"""
    ws = wb.create_sheet("MultiCurve_Framework")
    ws.append(["POST-2008 MULTI-CURVE FRAMEWORK"])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])
    
    # Introduction and explanation
    ws.append(["WHY MULTI-CURVE? (Post-Financial Crisis)"])
    ws['A3'].font = Font(bold=True, size=12, color="0066CC")
    ws.append([])
    
    explanation = [
        ("Before 2008:", "Single curve for discounting AND projection"),
        ("Problem:", "LIBOR-OIS basis became significant (credit/liquidity risk)"),
        ("Solution:", "Separate curves: OIS for discounting, LIBOR/SOFR for projection"),
        ("Benefit:", "More accurate swap pricing and risk management"),
        ("Standard:", "All major banks adopted this approach")
    ]
    
    for title, desc in explanation:
        ws.append([title, desc])
        if title.endswith(':'):
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="FF6600")
    
    ws.append([])
    
    # Curve Types and Usage
    ws.append(["CURVE TYPES IN MULTI-CURVE FRAMEWORK"])
    ws['A' + str(ws.max_row)].font = Font(bold=True, size=12, color="0066CC")
    ws.append([])
    
    headers = ["Curve Type", "Purpose", "Instruments", "Currency Examples"]
    ws.append(headers)
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=ws.max_row, column=i)
        cell.font = header_font
        cell.fill = header_fill
    
    curve_types = [
        ("OIS Discounting", "Present value calculation", "OIS Swaps", "Fed Funds, EONIA, SONIA"),
        ("LIBOR Projection", "Forward rate calculation", "LIBOR Swaps", "USD/EUR/GBP LIBOR"),
        ("SOFR Projection", "Forward rate (LIBOR replacement)", "SOFR Swaps", "USD SOFR"),
        ("EURIBOR Projection", "EUR forward rates", "EURIBOR Swaps", "EUR EURIBOR"),
        ("Government", "Risk-free benchmark", "Treasury Bonds", "UST, Bund, Gilt")
    ]
    
    for curve_type in curve_types:
        ws.append(curve_type)
        # Highlight the curves we're using
        if curve_type[0] in ["OIS Discounting", "SOFR Projection"]:
            ws.cell(row=ws.max_row, column=1).fill = highlight_fill
    
    ws.append([])
    
    # Current Implementation
    currency = bond_data.get('reference', {}).get('Position Currency', 'USD')
    ws.append([f"CURRENT IMPLEMENTATION ({currency})"])
    ws['A' + str(ws.max_row)].font = Font(bold=True, size=12, color="0066CC")
    ws.append([])
    
    # Determine curves based on currency
    if currency == 'USD':
        discounting_curve = "Fed Funds OIS"
        projection_curve = "SOFR (replacing USD LIBOR)"
        basis_description = "SOFR-Fed Funds Basis"
    elif currency == 'EUR':
        discounting_curve = "EONIA/€STR OIS" 
        projection_curve = "EURIBOR"
        basis_description = "EURIBOR-EONIA Basis"
    elif currency == 'GBP':
        discounting_curve = "SONIA OIS"
        projection_curve = "SONIA (replaced GBP LIBOR)"
        basis_description = "SONIA-Base Rate Basis"
    else:
        discounting_curve = "Government OIS"
        projection_curve = "Money Market Rate"
        basis_description = "Credit-Risk Free Basis"
    
    ws.append(["Discounting Curve:", discounting_curve, "Used for present value calculations", "Lower rates (risk-free)"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=2).fill = input_fill
    
    ws.append(["Projection Curve:", projection_curve, "Used for forward rate calculations", "Higher rates (credit premium)"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=2).fill = input_fill
    
    ws.append(["Basis Spread:", basis_description, "Difference between the two curves", "Reflects credit/liquidity risk"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=2).fill = highlight_fill
    
    ws.append([])
    
    # Demonstration with Current Curve Data
    ws.append(["SINGLE vs MULTI-CURVE COMPARISON"])
    ws['A' + str(ws.max_row)].font = Font(bold=True, size=12, color="0066CC")
    ws.append([])
    
    headers = ["Maturity", "Single Curve Rate", "OIS Discounting", "Projection Rate", "Basis (bps)", "Impact"]
    ws.append(headers)
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=ws.max_row, column=i)
        cell.font = header_font
        cell.fill = header_fill
    
    # Get curve data
    times, rates = curve_data
    
    # Create synthetic multi-curve data (in practice would be bootstrapped)
    for i, (time, rate) in enumerate(zip(times[:5], rates[:5])):  # First 5 points
        # Synthetic OIS curve (lower by 10-50 bps depending on maturity)
        if time <= 1:
            ois_adjustment = 0.0010  # 10 bps for short end
        elif time <= 2:
            ois_adjustment = 0.0025  # 25 bps
        else:
            ois_adjustment = 0.0035  # 35 bps for longer end
            
        single_curve_rate = rate
        ois_rate = rate - ois_adjustment
        projection_rate = rate  # Keep original as projection
        basis_bps = (projection_rate - ois_rate) * 10000
        
        # Format time as tenor
        if time < 1:
            tenor = f"{int(time*12)}M"
        else:
            tenor = f"{int(time)}Y"
        
        # Calculate row number before adding to avoid circular references
        next_row = ws.max_row + 1
        
        ws.append([
            tenor,
            f"{single_curve_rate:.4f}",
            f"=B{next_row}*{1-ois_adjustment/rate:.6f}",
            f"=B{next_row}",
            f"=(D{next_row}-C{next_row})*10000", 
            f"{basis_bps:.0f} bps difference"
        ])
        
        # Add formula fill for calculated cells
        row_num = ws.max_row
        ws.cell(row=row_num, column=3).fill = formula_fill
        ws.cell(row=row_num, column=4).fill = formula_fill  
        ws.cell(row=row_num, column=5).fill = formula_fill
        
        # Highlight basis if significant
        if basis_bps > 20:
            ws.cell(row=row_num, column=5).fill = highlight_fill
    
    ws.append([])
    
    # Swap Pricing Example
    ws.append(["SWAP PRICING WITH MULTI-CURVE"])
    ws['A' + str(ws.max_row)].font = Font(bold=True, size=12, color="0066CC")
    ws.append([])
    
    ws.append(["Example: 5-Year Interest Rate Swap"])
    ws.append([])
    
    # Swap parameters
    notional = 1000000
    swap_rate = 0.045  # 4.5%
    
    headers = ["Payment", "Time", "Fixed Rate", "Projection Rate", "Fixed PV", "Float PV", "Net PV"]
    ws.append(headers)
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=ws.max_row, column=i)
        cell.font = header_font
        cell.fill = header_fill
    
    # Simplified swap cashflows (semiannual)
    swap_start_row = ws.max_row + 1
    for payment in range(1, 11):  # 10 semiannual payments
        time_years = payment * 0.5
        
        # Get rates (interpolated from our curve data)
        if time_years <= max(times):
            # Simple interpolation
            rate_index = min(int(time_years), len(rates)-1)
            base_rate = rates[rate_index] if rate_index < len(rates) else rates[-1]
        else:
            base_rate = rates[-1]
        
        # Multi-curve rates
        ois_rate = base_rate - 0.0030  # 30 bps lower for discounting
        projection_rate = base_rate     # Use for forward rates
        
        # Calculate row number before adding to avoid circular references
        next_row = ws.max_row + 1
        
        ws.append([
            payment,
            time_years,
            swap_rate,
            projection_rate,
            f"={swap_rate/2}*{notional/100}*EXP(-C{next_row}*B{next_row})",  # Fixed leg PV using OIS
            f"={projection_rate/2}*{notional/100}*EXP(-{ois_rate}*B{next_row})",  # Float leg PV
            f"=E{next_row}-F{next_row}"  # Net PV
        ])
        
        # Formula fill
        row_num = ws.max_row
        for col in range(5, 8):
            ws.cell(row=row_num, column=col).fill = formula_fill
    
    swap_end_row = ws.max_row
    
    # Sum totals
    ws.append([])
    ws.append(["", "", "TOTALS:", "", f"=SUM(E{swap_start_row}:E{swap_end_row})", 
              f"=SUM(F{swap_start_row}:F{swap_end_row})", f"=SUM(G{swap_start_row}:G{swap_end_row})"])
    total_row = ws.max_row
    
    for col in range(5, 8):
        ws.cell(row=total_row, column=col).fill = highlight_fill
        ws.cell(row=total_row, column=col).font = Font(bold=True)
    
    ws.append([])
    
    # Basis Spread Analysis
    ws.append(["BASIS SPREAD TERM STRUCTURE"])
    ws['A' + str(ws.max_row)].font = Font(bold=True, size=12, color="0066CC")
    ws.append([])
    
    ws.append(["Shows credit/liquidity premium over risk-free rate"])
    ws.append([])
    
    headers = ["Tenor", "Risk-Free (OIS)", "Credit Rate", "Basis Spread", "Basis (bps)", "Interpretation"]
    ws.append(headers)
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=ws.max_row, column=i)
        cell.font = header_font
        cell.fill = header_fill
    
    # Term structure of basis spreads
    basis_tenors = ["3M", "6M", "1Y", "2Y", "5Y", "10Y"]
    basis_values = [15, 20, 25, 30, 35, 30]  # Typical basis spread pattern
    
    for i, (tenor, basis_bps) in enumerate(zip(basis_tenors, basis_values)):
        if i < len(rates):
            credit_rate = rates[min(i, len(rates)-1)]
            ois_rate = credit_rate - (basis_bps / 10000)
        else:
            credit_rate = 0.05
            ois_rate = credit_rate - (basis_bps / 10000)
        
        if basis_bps <= 20:
            interpretation = "Low credit risk"
        elif basis_bps <= 35:
            interpretation = "Moderate risk"
        else:
            interpretation = "Elevated risk"
        
        # Calculate row number before adding to avoid circular references
        next_row = ws.max_row + 1
        
        ws.append([
            tenor,
            f"{ois_rate:.4f}",
            f"{credit_rate:.4f}", 
            f"=C{next_row}-B{next_row}",
            f"=D{next_row}*10000",
            interpretation
        ])
        
        # Formula fills
        row_num = ws.max_row
        ws.cell(row=row_num, column=4).fill = formula_fill
        ws.cell(row=row_num, column=5).fill = formula_fill
        
        # Color code basis spread
        if basis_bps > 30:
            ws.cell(row=row_num, column=5).fill = highlight_fill
    
    ws.append([])
    
    # Implementation Notes  
    ws.append(["IMPLEMENTATION NOTES"])
    ws['A' + str(ws.max_row)].font = Font(bold=True, size=12, color="0066CC")
    ws.append([])
    
    implementation_notes = [
        ("Curve Construction:", "Bootstrap OIS and projection curves simultaneously"),
        ("Basis Calibration:", "Use basis swaps to calibrate spread relationships"),
        ("Interpolation:", "Maintain smoothness while preserving market relationships"),
        ("Risk Management:", "Calculate sensitivities to both curves separately"),
        ("Regulatory:", "Basel III requires multi-curve for derivative valuation"),
        ("Systems Impact:", "Requires significant changes to pricing systems")
    ]
    
    for note_title, note_desc in implementation_notes:
        ws.append([note_title, note_desc])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="FF6600")
    
    ws.append([])
    
    # Connection to Current Bond
    ws.append(["APPLICATION TO CURRENT BOND"])
    ws['A' + str(ws.max_row)].font = Font(bold=True, size=12, color="0066CC")
    ws.append([])
    
    # Show impact on current bond calculation
    single_curve_ytm = python_results.get('ytm', 0.05)
    multi_curve_adjustment = 0.0015  # Typical 15 bps adjustment
    
    ws.append(["Single Curve YTM:", f"{single_curve_ytm:.4f}", "Traditional calculation", "Uses one curve for everything"])
    single_ytm_row = ws.max_row
    ws.append(["Multi-Curve Adjustment:", f"{multi_curve_adjustment:.4f}", "Credit/liquidity premium", "Basis spread impact"])
    adjustment_row = ws.max_row
    ws.append(["Adjusted YTM:", f"=B{single_ytm_row}+B{adjustment_row}", "More accurate measure", "Reflects true funding cost"])
    ws.cell(row=ws.max_row, column=2).fill = formula_fill
    ws.cell(row=ws.max_row, column=2).fill = highlight_fill
    
    # Impact on price
    duration = python_results.get('modified_duration', 5.0)
    price_impact = duration * multi_curve_adjustment * 100  # Price impact per $100
    
    ws.append([])
    ws.append(["Price Impact Analysis:"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="FF6600")
    ws.append(["Modified Duration:", f"{duration:.2f}", "From main calculation", "Years"])
    duration_row = ws.max_row
    # Fix: Reference the specific adjustment row and duration row
    ws.append(["Basis Impact on Price:", f"=B{adjustment_row}*B{duration_row}*100", f"${price_impact:.4f} per $100", "Duration × basis × 100"])
    ws.cell(row=ws.max_row, column=2).fill = formula_fill
    ws.cell(row=ws.max_row, column=2).fill = highlight_fill
    
    # Format columns
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18  
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 25
