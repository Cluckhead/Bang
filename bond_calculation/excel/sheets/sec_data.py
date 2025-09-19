# sheets/sec_data.py
# Purpose: Build the Sec_Data sheet displaying security time series data

from __future__ import annotations

from datetime import datetime
from typing import Dict, List
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis

from ..styles import header_font, header_fill


def add_sec_data_sheet(
    wb: Workbook,
    bond_data: Dict,
    valuation_date: datetime,
    sec_data: Dict = None
) -> None:
    """
    Create a Sec_Data sheet displaying security-level time series data.
    Shows latest snapshot and recent history from sec_*.csv files.
    """
    ws = wb.create_sheet("Sec_Data")
    
    # Title
    ws.merge_cells('A1:I1')
    ws['A1'] = f"SECURITY TIME SERIES DATA - {bond_data['reference']['ISIN']}"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Section styling
    section_font = Font(bold=True, size=11, color="FFFFFF")
    section_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    
    row = 3
    
    # Latest Snapshot Section
    ws[f'A{row}'] = "LATEST SNAPSHOT"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:I{row}')
    row += 1
    
    # Snapshot headers
    snapshot_headers = ["Metric", "Value", "Date", "Source", "Change", "% Change"]
    for col, header in enumerate(snapshot_headers, 1):
        ws.cell(row, col, header)
        ws.cell(row, col).font = header_font
        ws.cell(row, col).fill = header_fill
    row += 1
    
    # Sample snapshot data (would be populated from sec_*.csv files)
    if sec_data and 'snapshot' in sec_data:
        snapshot = sec_data['snapshot']
    else:
        # Default sample data
        snapshot = {
            'Price': {'value': 98.5, 'date': valuation_date.strftime('%d/%m/%Y'), 'change': -0.25, 'pct_change': -0.253},
            'Spread (bps)': {'value': 125, 'date': valuation_date.strftime('%d/%m/%Y'), 'change': 2, 'pct_change': 1.626},
            'YTM (%)': {'value': 4.75, 'date': valuation_date.strftime('%d/%m/%Y'), 'change': 0.05, 'pct_change': 1.064},
            'YTW (%)': {'value': 4.60, 'date': valuation_date.strftime('%d/%m/%Y'), 'change': 0.03, 'pct_change': 0.656},
            'Duration': {'value': 7.25, 'date': valuation_date.strftime('%d/%m/%Y'), 'change': -0.02, 'pct_change': -0.275},
            'Convexity': {'value': 62.3, 'date': valuation_date.strftime('%d/%m/%Y'), 'change': 0.1, 'pct_change': 0.161},
        }
    
    for metric, data in snapshot.items():
        ws[f'A{row}'] = metric
        ws[f'B{row}'] = float(data['value']) if data['value'] is not None else ''
        ws[f'C{row}'] = data['date']
        ws[f'D{row}'] = f"sec_{metric.split()[0].lower()}.csv"
        
        change_val = data.get('change')
        if change_val is not None and change_val != '':
            ws[f'E{row}'] = float(change_val)
        else:
            ws[f'E{row}'] = ''
            
        pct_val = data.get('pct_change')
        if pct_val is not None and pct_val != '':
            ws[f'F{row}'] = float(pct_val)
            ws[f'F{row}'].number_format = '0.00%'
        else:
            ws[f'F{row}'] = ''
        
        # Color coding for changes
        if data.get('change'):
            change_cell = ws[f'E{row}']
            pct_cell = ws[f'F{row}']
            if data['change'] > 0:
                change_cell.font = Font(color="27AE60")  # Green for positive
                pct_cell.font = Font(color="27AE60")
            elif data['change'] < 0:
                change_cell.font = Font(color="E74C3C")  # Red for negative
                pct_cell.font = Font(color="E74C3C")
        row += 1
    
    row += 2
    
    # Time Series Section
    ws[f'A{row}'] = "RECENT TIME SERIES (LAST 20 DAYS)"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:I{row}')
    row += 1
    
    # Time series headers
    ts_headers = ["Date", "Price", "Spread", "YTM", "YTW", "Duration", "DTS", "Convexity", "Source"]
    for col, header in enumerate(ts_headers, 1):
        ws.cell(row, col, header)
        ws.cell(row, col).font = header_font
        ws.cell(row, col).fill = header_fill
    row += 1
    
    ts_start_row = row
    
    # Sample time series data (would be populated from sec_*.csv files)
    if sec_data and 'time_series' in sec_data:
        time_series = sec_data['time_series'][-20:]  # Last 20 days
    else:
        # Generate sample declining time series
        import random
        time_series = []
        for i in range(20, 0, -1):
            date = valuation_date - pd.Timedelta(days=i)
            time_series.append({
                'date': date.strftime('%d/%m/%Y'),
                'price': 98.5 + random.uniform(-0.5, 0.5),
                'spread': 125 + random.randint(-5, 5),
                'ytm': 4.75 + random.uniform(-0.1, 0.1),
                'ytw': 4.60 + random.uniform(-0.1, 0.1),
                'duration': 7.25 + random.uniform(-0.05, 0.05),
                'dts': 0.01 + random.uniform(-0.005, 0.005),
                'convexity': 62.3 + random.uniform(-0.5, 0.5),
            })
    
    for ts_point in time_series:
        ws[f'A{row}'] = ts_point.get('date', '')
        
        # Save numeric values as numbers, not formatted strings
        price_val = ts_point.get('price')
        if price_val is not None and price_val != '':
            ws[f'B{row}'] = float(price_val)
            ws[f'B{row}'].number_format = '0.000'
        else:
            ws[f'B{row}'] = ''
            
        spread_val = ts_point.get('spread')
        if spread_val is not None and spread_val != '':
            ws[f'C{row}'] = float(spread_val)
        else:
            ws[f'C{row}'] = ''
            
        ytm_val = ts_point.get('ytm')
        if ytm_val is not None and ytm_val != '':
            ws[f'D{row}'] = float(ytm_val)
            ws[f'D{row}'].number_format = '0.000'
        else:
            ws[f'D{row}'] = ''
            
        ytw_val = ts_point.get('ytw')
        if ytw_val is not None and ytw_val != '':
            ws[f'E{row}'] = float(ytw_val)
            ws[f'E{row}'].number_format = '0.000'
        else:
            ws[f'E{row}'] = ''
            
        duration_val = ts_point.get('duration')
        if duration_val is not None and duration_val != '':
            ws[f'F{row}'] = float(duration_val)
            ws[f'F{row}'].number_format = '0.000'
        else:
            ws[f'F{row}'] = ''
            
        dts_val = ts_point.get('dts')
        if dts_val is not None and dts_val != '':
            ws[f'G{row}'] = float(dts_val)
            ws[f'G{row}'].number_format = '0.0000'
        else:
            ws[f'G{row}'] = ''
            
        convexity_val = ts_point.get('convexity')
        if convexity_val is not None and convexity_val != '':
            ws[f'H{row}'] = float(convexity_val)
            ws[f'H{row}'].number_format = '0.00'
        else:
            ws[f'H{row}'] = ''
            
        ws[f'I{row}'] = "sec_*.csv"
        row += 1
    
    ts_end_row = row - 1
    
    # Define named range for time series data
    if ts_end_row >= ts_start_row:
        wb.defined_names.add(DefinedName("Sec_TimeSeries", 
                                         attr_text=f"Sec_Data!$A${ts_start_row}:$I${ts_end_row}"))
    
    # Add a simple chart for Price and Spread
    if ts_end_row >= ts_start_row + 5:  # Only add chart if we have enough data points
        row += 2
        ws[f'A{row}'] = "PRICE & SPREAD TREND"
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        ws.merge_cells(f'A{row}:I{row}')
        
        # Create line chart
        chart = LineChart()
        chart.title = "Price and Spread Over Time"
        chart.style = 12
        chart.height = 10
        chart.width = 20
        chart.y_axis.title = "Price"
        chart.x_axis.title = "Date"
        
        # Add price data
        price_data = Reference(ws, min_col=2, min_row=ts_start_row-1, 
                               max_col=2, max_row=ts_end_row)
        dates = Reference(ws, min_col=1, min_row=ts_start_row, 
                         max_row=ts_end_row)
        chart.add_data(price_data, titles_from_data=True)
        chart.set_categories(dates)
        
        # Position chart
        chart_row = row + 2
        chart.anchor = f'A{chart_row}'
        ws.add_chart(chart)
    
    row += 20  # Space for chart
    
    # Statistics Section
    ws[f'A{row}'] = "STATISTICS (LAST 20 DAYS)"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:I{row}')
    row += 1
    
    stats_headers = ["Metric", "Mean", "Std Dev", "Min", "Max", "Current", "Z-Score"]
    for col, header in enumerate(stats_headers, 1):
        ws.cell(row, col, header)
        ws.cell(row, col).font = header_font
        ws.cell(row, col).fill = header_fill
    row += 1
    
    # Calculate statistics using Excel formulas
    metrics_to_analyze = [
        ("Price", 2),
        ("Spread", 3),
        ("YTM", 4),
        ("Duration", 6),
    ]
    
    for metric_name, col_num in metrics_to_analyze:
        col_letter = get_column_letter(col_num)
        ws[f'A{row}'] = metric_name
        ws[f'B{row}'] = f"=AVERAGE({col_letter}{ts_start_row}:{col_letter}{ts_end_row})"
        ws[f'C{row}'] = f"=STDEV({col_letter}{ts_start_row}:{col_letter}{ts_end_row})"
        ws[f'D{row}'] = f"=MIN({col_letter}{ts_start_row}:{col_letter}{ts_end_row})"
        ws[f'E{row}'] = f"=MAX({col_letter}{ts_start_row}:{col_letter}{ts_end_row})"
        ws[f'F{row}'] = f"={col_letter}{ts_end_row}"  # Current (last) value
        ws[f'G{row}'] = f"=IF(C{row}>0,(F{row}-B{row})/C{row},0)"  # Z-Score
        
        # Format Z-Score cell
        z_cell = ws[f'G{row}']
        z_cell.number_format = '0.00'
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 15
    
    # Freeze panes
    ws.freeze_panes = 'A5'