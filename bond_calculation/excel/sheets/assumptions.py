# sheets/assumptions.py
# Purpose: Build the Assumptions sheet and define named variables used by other sheets

from __future__ import annotations

from typing import Dict

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.workbook.defined_name import DefinedName

from ..styles import header_font, header_fill, input_fill


def _add_defined_name_safe(wb: Workbook, name: str, attr_text: str) -> None:
    try:
        dn = DefinedName(name=name, attr_text=attr_text)
        # openpyxl has both add/append across versions
        if hasattr(wb.defined_names, "add"):
            wb.defined_names.add(dn)
        elif hasattr(wb.defined_names, "append"):
            wb.defined_names.append(dn)
    except Exception:
        # best-effort; non-fatal
        pass


def add_assumptions_sheet(wb: Workbook, bond_data: Dict) -> None:
    ws = wb.create_sheet("Assumptions")
    ws.append(["SCENARIO ASSUMPTIONS (Named)"])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Name", "Value", "Description"])
    for i in range(1, 4):
        cell = ws.cell(row=3, column=i)
        cell.font = header_font
        cell.fill = header_fill

    # Values sourced from schedule/defaults
    coupon_freq = int(bond_data['schedule'].get('Coupon Frequency', 2))
    day_basis = str(bond_data['schedule'].get('Day Basis', 'ACT/ACT'))

    assumptions_list = [
        ("Assump_Frequency", coupon_freq, "Coupon payments per year"),
        ("Assump_Basis", day_basis, "Day count basis (display)"),
        ("Assump_CurveShift_bps", 0, "Parallel curve shift (bps)"),
        ("Assump_ZShift_bps", 0, "Additional z-spread shift (bps)"),
        ("Assump_InterpMethod", "Linear", "Interpolation: Linear or Flat"),
        ("Assump_Notional", 100, "Face value"),
        ("Assump_PV_Tolerance", 0.0001, "PV convergence tolerance"),
        ("Assump_Max_Iterations", 100, "Solver max iterations"),
        ("Assump_Default_Bumps", 1, "Default bump size (bps)"),
        ("Assump_Gaussian_Width", 2, "KRD Gaussian width (years)"),
        ("Assump_Settlement_Days", 2, "Default settlement days"),
    ]

    for name, value, desc in assumptions_list:
        ws.append([name.replace('Assump_', ''), value, desc])
        current_row = ws.max_row
        _add_defined_name_safe(wb, name, f"Assumptions!$B${current_row}")
        if name not in ("Assump_Basis", "Assump_InterpMethod"):
            ws.cell(row=current_row, column=2).fill = input_fill

    # Basis Code (numeric) derived from Assump_Basis for Excel functions expecting a basis enum
    ws.append(["Basis Code", None, "Numeric basis for Excel functions"])
    basis_row = ws.max_row
    ws.cell(row=basis_row, column=2).value = (
        '=IF(UPPER(Assump_Basis)="ACT/ACT",1,'
        'IF(UPPER(Assump_Basis)="ACT/360",2,'
        'IF(UPPER(Assump_Basis)="ACT/365",3,'
        'IF(UPPER(Assump_Basis)="30E/360",4,'
        'IF(UPPER(Assump_Basis)="30/360",0,1)))))'
    )
    _add_defined_name_safe(wb, 'Assump_Basis_Code', f"Assumptions!$B${basis_row}")


