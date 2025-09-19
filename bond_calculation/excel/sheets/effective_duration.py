# sheets/effective_duration.py
# Purpose: Build the Effective_Duration sheet with Python results and formula demo

from __future__ import annotations

from typing import Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font
from ..styles import header_font, header_fill, input_fill


def add_effective_duration_sheet(wb: Workbook, python_results: Dict, cashflows: List[Dict]) -> None:
    ws = wb.create_sheet("Effective_Duration")
    ws.append(["EFFECTIVE DURATION CALCULATION"])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Python Calculated Results:"])
    ws['A3'].font = Font(bold=True, color="0066CC")
    ws.append(["Effective Duration:", python_results.get('effective_duration', 0.0), "years"])
    ws.append(["Modified Duration:", python_results.get('modified_duration', 0.0), "years"])
    ws.append(["Convexity:", python_results.get('convexity', 0.0), ""])
    ws.append(["Spread Duration:", python_results.get('spread_duration', 0.0), "years"])
    ws.append([])

    ws.append(["Excel Formula Calculation:"])
    ws['A9'].font = Font(bold=True, color="0066CC")
    ws.append(["Rate Shock (bps):", 10, "← Modify to see impact"])
    ws.cell(row=10, column=2).fill = input_fill
    ws.append([])

    headers = ["CF #", "Time", "Cashflow", "Base Zero", "Zero+Shock", "Zero-Shock", "PV Base", "PV Up", "PV Down"]
    ws.append(headers)
    for i in range(1, len(headers) + 1):
        cell = ws.cell(row=12, column=i)
        cell.font = header_font
        cell.fill = header_fill

    start_row = 13
    max_rows = max(1, len(cashflows))
    for i in range(1, max_rows + 1):
        r = start_row + i - 1
        ws.append([
            i,
            f"=Cashflows!D{i+1}",
            f"=Cashflows!I{i+1}",
            f"=FORECAST(B{r},Curve_Rates,Curve_Terms)/100",
            f"=D{r}+$B$10/10000",
            f"=D{r}-$B$10/10000",
            f"=C{r}/(1+D{r}/Assump_Frequency)^(Assump_Frequency*B{r})",
            f"=C{r}/(1+E{r}/Assump_Frequency)^(Assump_Frequency*B{r})",
            f"=C{r}/(1+F{r}/Assump_Frequency)^(Assump_Frequency*B{r})",
        ])

    last_cf_row = start_row + max_rows - 1
    ws.append([])
    ws.append(["", "", "Total PV:", "", "", "", f"=SUM(G{start_row}:G{last_cf_row})", f"=SUM(H{start_row}:H{last_cf_row})", f"=SUM(I{start_row}:I{last_cf_row})"])
    total_row = ws.max_row
    ws.append([])
    # Provide a working Excel formula and a plain-text explanation. Avoid @-style references.
    ws.append(["Effective Duration Formula:", f"=(I{total_row}-H{total_row})/(2*G{total_row}*$B$10/10000)", "(PV_down - PV_up) / (2 * Price * Shock)"])
    ws.cell(row=ws.max_row, column=2).number_format = '0.0000'
    ws.append(["Modified Duration Formula:", f"=B{ws.max_row}/(1+YTM_Calculations!$B$5/Assump_Frequency)", "Eff_Duration / (1 + YTM/Frequency)"])
    ws.cell(row=ws.max_row, column=2).number_format = '0.0000'


