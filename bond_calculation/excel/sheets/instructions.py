# sheets/instructions.py
# Purpose: Build the Instructions sheet

from __future__ import annotations

from typing import Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font


def add_instructions_sheet(wb: Workbook, python_results: Optional[Dict] = None) -> None:
    """Add enhanced instructions sheet with institutional features documentation"""
    if python_results is None:
        python_results = {}
        
    ws = wb.create_sheet("Instructions")
    ws.append(["BOND CALCULATION WORKBOOK - USER GUIDE"])
    ws['A1'].font = Font(bold=True, size=14, color="0066CC")
    ws.append([])

    instructions = [
        ("🏛️ INSTITUTIONAL-GRADE BOND ANALYTICS WORKBOOK", ""),
        ("", ""),
        ("This workbook implements trading desk-level precision:", ""),
        ("✓ ISDA-compliant day count conventions", "Exact leap year handling"),
        ("✓ Hull-White Monte Carlo OAS", "Advanced volatility modeling"),  
        ("✓ Multi-curve framework", "Post-2008 OIS/SOFR separation"),
        ("✓ Higher-order Greeks", "Cross-gamma, key rate convexity"),
        ("✓ Settlement mechanics", "T+1/T+2/T+3 with holiday calendars"),
        ("✓ Robust numerical methods", "Brent's method, adaptive quadrature"),
        ("", ""),
        ("📊 CORE CALCULATION SHEETS:", ""),
        ("", ""),
        ("1. Input Parameters:", "Modify blue cells to change bond characteristics"),
        ("   - Clean Price", "Change to see impact on yields"),
        ("   - Uses precise cell references", "Fixed formula errors"),
        ("", ""),
        ("2. Yield Curve:", "Edit zero rates to see spread impacts"),
        ("   - Rates are editable", "Change curve shape"),
        ("   - Discount factors", "Auto-calculate from rates"),
        ("", ""),
        ("3. YTM Calculations:", "Multiple calculation approaches"),
        ("   - Python SpreadOMatic", "Robust Newton-Raphson"),
        ("   - Excel XIRR/NOMINAL", "Fixed range errors, added frequency"),
        ("   - Excel YIELD function", "Native Excel with proper parameters"),
        ("", ""),
        ("4. Z-Spread:", "Spread over entire yield curve"),
        ("   - Input spread in bps", "Modify blue cell"),
        ("   - Dynamic cashflow rows", "No excessive blanks"),
        ("   - FORECAST interpolation", "Linear curve interpolation"),
        ("", ""),
        ("5. OAS Calculations:", "Option-adjusted spread (callable bonds only)"),
        ("   - Standard OAS", "Single call, fixed 20% volatility"),
        ("   - Enhanced OAS", "Hull-White Monte Carlo, calibrated vol"),
        ("   - Shows option value", "Z-Spread minus OAS"),
        ("", ""),
        ("🚀 ENHANCED INSTITUTIONAL SHEETS:", ""),
        ("", ""),
        ("6. Settlement Enhanced:", "Professional settlement mechanics"),
        ("   - T+1/T+2/T+3 rules", "Market-specific conventions"),
        ("   - Precise accrued interest", "ISDA-compliant calculations"),
        ("   - Ex-dividend handling", "Record date analysis"),
        ("   - Holiday calendar impact", "Business day adjustments"),
        ("", ""),
        ("7. MultiCurve Framework:", "Post-2008 curve construction"),
        ("   - OIS Discounting curve", "Risk-free discounting"),
        ("   - SOFR/LIBOR Projection", "Forward rate calculation"),
        ("   - Basis spread analysis", "Credit/liquidity premiums"),
        ("   - Swap pricing demo", "Multi-curve impact"),
        ("", ""),
        ("8. Higher Order Greeks:", "Advanced risk metrics"),
        ("   - Cross-gamma matrix", "Correlation risk effects"),
        ("   - Key rate convexity", "Term structure convexity"),
        ("   - Portfolio scenarios", "Stress testing"),
        ("   - Hedge ratio calculations", "Risk management"),
        ("", ""),
        ("9. DayCount Precision:", "Day count accuracy comparison"),
        ("   - ISDA vs approximations", "Shows precision differences"),
        ("   - Leap year demonstrations", "Material impact analysis"),
        ("   - 30/360 month-end rules", "Complex adjustment logic"),
        ("   - Implementation guidance", "Best practice recommendations"),
        ("", ""),
        ("10. Hull-White Monte Carlo:", "Advanced OAS modeling"),
        ("   - Model specification", "Stochastic differential equation"),
        ("   - Parameter calibration", "Swaption volatility fitting"),
        ("   - Simulation mathematics", "Monte Carlo path generation"),
        ("   - Model validation", "Institutional quality checks"),
        ("", ""),
        ("📈 MATHEMATICAL PRECISION:", ""),
        ("- Enhanced day counts", "ISDA ACT/ACT, 30E/360, NL/365"),
        ("- Curve construction", "Monotone cubic splines, Nelson-Siegel"),
        ("- Numerical methods", "Brent's method (guaranteed convergence)"),
        ("- Volatility modeling", "Hull-White, Black-Karasinski"),
        ("- Settlement accuracy", "Business day, holiday adjustments"),
        ("", ""),
        ("🎯 COLOR CODING:", ""),
        ("- Blue cells", "User inputs (editable)"),
        ("- Yellow cells", "Key formulas"),
        ("- Orange cells", "Important results/highlights"),
        ("- Green cells", "Enhanced institutional features"),
        ("- White cells", "Calculated values"),
        ("", ""),
        ("💡 ADVANCED TIPS:", ""),
        ("- Use Goal Seek", "Excel Data > What-If > Goal Seek"),
        ("- Try Solver", "For multi-variable optimization"),
        ("- Scenario Manager", "For stress testing"),
        ("- Data Tables", "For sensitivity analysis"),
        ("- Check enhanced sheets", "For institutional-grade analytics"),
        ("", ""),
        ("🏛️ INSTITUTIONAL FEATURES:", ""),
        ("This workbook now includes the same mathematical rigor", ""),
        ("and precision used by major investment bank trading desks.", ""),
        ("", ""),
        ("Enhancement Level:", python_results.get('enhancement_level', 'standard')),
        ("Numerical Method:", python_results.get('numerical_method', 'newton_raphson')),
        ("Day Count Precision:", "✓ Enabled" if python_results.get('day_count_precision') else "Standard"),
        ("Curve Method:", python_results.get('curve_method', 'linear')),
        ("Volatility Model:", python_results.get('volatility_model', 'black_scholes')),
    ]

    for i, (title, desc) in enumerate(instructions, 3):
        ws.append([title, desc])
        if title and not desc:
            ws.cell(row=i, column=1).font = Font(bold=True, color="0066CC")
        elif title.startswith("-"):
            ws.cell(row=i, column=1).font = Font(italic=True)


