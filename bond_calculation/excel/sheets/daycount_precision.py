# sheets/daycount_precision.py
# Purpose: Day count precision sheet showing exact ISDA calculations vs approximations
# Demonstrates the accuracy improvements from institutional-grade day count methods

from __future__ import annotations

from datetime import datetime, timedelta
from typing import Dict

from openpyxl import Workbook
from openpyxl.styles import Font
from ..styles import header_font, header_fill, input_fill, formula_fill, highlight_fill


def add_daycount_precision_sheet(wb: Workbook, bond_data: Dict, valuation_date: datetime) -> None:
    """Add day count precision sheet showing institutional vs approximate methods"""
    ws = wb.create_sheet("DayCount_Precision")
    ws.append(["DAY COUNT CONVENTION PRECISION"])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])
    
    # Introduction
    ws.append(["WHY DAY COUNT PRECISION MATTERS"])
    ws['A3'].font = Font(bold=True, size=12, color="0066CC")
    ws.append([])
    
    intro_points = [
        ("Accrued Interest:", "Small day count errors compound in portfolio calculations"),
        ("ISDA Standards:", "Legal framework requires exact day count methods"),
        ("Leap Years:", "ACT/365 vs ACT/365.25 vs ACT/ACT creates material differences"),
        ("Month Ends:", "30/360 has complex rules for month-end dates"),
        ("Cross-Border:", "Different markets use different conventions")
    ]
    
    for title, desc in intro_points:
        ws.append([title, desc])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="FF6600")
    
    ws.append([])
    
    # Day Count Convention Reference
    ws.append(["DAY COUNT CONVENTION REFERENCE"])
    ws['A' + str(ws.max_row)].font = Font(bold=True, size=12, color="0066CC")
    ws.append([])
    
    headers = ["Convention", "Numerator", "Denominator", "Usage", "Complexity"]
    ws.append(headers)
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=ws.max_row, column=i)
        cell.font = header_font
        cell.fill = header_fill
    
    conventions = [
        ("30/360 (Bond)", "30-day months", "360", "US Corporate bonds", "Medium"),
        ("30E/360 (European)", "European 30/360", "360", "European bonds", "Medium"),
        ("ACT/360", "Actual days", "360", "Money markets", "Simple"),
        ("ACT/365 Fixed", "Actual days", "365", "Some bonds", "Simple"),
        ("ACT/365.25", "Actual days", "365.25", "Approximation", "Simple"),
        ("ACT/ACT ISDA", "Actual by year", "Actual by year", "ISDA standard", "Complex"),
        ("ACT/ACT ICMA", "Actual by period", "Actual by period", "ICMA bonds", "Complex"),
        ("NL/365", "No leap days", "365", "Specialized", "Medium")
    ]
    
    for convention in conventions:
        ws.append(convention)
        # Highlight the most precise methods
        if "ISDA" in convention[0] or "ICMA" in convention[0]:
            ws.cell(row=ws.max_row, column=1).fill = highlight_fill
        # Show complexity
        if convention[4] == "Complex":
            ws.cell(row=ws.max_row, column=5).fill = highlight_fill
    
    ws.append([])
    
    # Precision Comparison Using Current Bond
    bond_day_count = bond_data.get('schedule', {}).get('Day Basis', 'ACT/ACT')
    coupon_rate = float(bond_data.get('reference', {}).get('Coupon Rate', 5.0))
    
    ws.append([f"PRECISION COMPARISON FOR CURRENT BOND ({bond_day_count})"])
    ws['A' + str(ws.max_row)].font = Font(bold=True, size=12, color="0066CC")
    ws.append([])
    
    # Test dates - include leap year scenarios
    test_scenarios = [
        ("Regular Period", datetime(2024, 1, 15), datetime(2024, 7, 15), "6-month regular period"),
        ("Leap Year Feb", datetime(2024, 1, 29), datetime(2024, 2, 29), "Includes Feb 29 (leap year)"),
        ("Non-Leap Feb", datetime(2023, 1, 28), datetime(2023, 2, 28), "Non-leap year February"),
        ("Month End", datetime(2024, 1, 31), datetime(2024, 2, 29), "31st to month end"),
        ("Year Boundary", datetime(2023, 12, 15), datetime(2024, 1, 15), "Crosses year boundary")
    ]
    
    for scenario_name, start_date, end_date, description in test_scenarios:
        ws.append([f"SCENARIO: {scenario_name} ({description})"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="FF6600")
        ws.append([])
        
        # Date range
        ws.append(["Start Date:", start_date.strftime('%d/%m/%Y'), start_date.strftime('%A')])
        ws.append(["End Date:", end_date.strftime('%d/%m/%Y'), end_date.strftime('%A')])
        ws.append(["Calendar Days:", f"{(end_date - start_date).days}", "Simple difference"])
        ws.append([])
        
        # Comparison table
        headers = ["Method", "Day Count Fraction", "Formula/Logic", "Accrued Interest", "Error vs ISDA"]
        ws.append(headers)
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=ws.max_row, column=i)
            cell.font = header_font
            cell.fill = header_fill
        
        # Calculate different methods
        calendar_days = (end_date - start_date).days
        
        # 1. ACT/365 Fixed (simple)
        act_365_fraction = calendar_days / 365.0
        act_365_accrued = act_365_fraction * (coupon_rate/100) / 2  # Assume semiannual
        
        # 2. ACT/365.25 (approximation)
        act_365_25_fraction = calendar_days / 365.25
        act_365_25_accrued = act_365_25_fraction * (coupon_rate/100) / 2
        
        # 3. ACT/360 (money market)
        act_360_fraction = calendar_days / 360.0
        act_360_accrued = act_360_fraction * (coupon_rate/100) / 2
        
        # 4. 30/360 (simplified calculation)
        # Simplified 30/360 - real implementation is more complex
        if start_date.day == 31:
            adj_start_day = 30
        else:
            adj_start_day = start_date.day
        
        if end_date.day == 31 and adj_start_day == 30:
            adj_end_day = 30
        else:
            adj_end_day = end_date.day
        
        days_30_360 = (end_date.year - start_date.year) * 360 + \
                     (end_date.month - start_date.month) * 30 + \
                     (adj_end_day - adj_start_day)
        
        fraction_30_360 = days_30_360 / 360.0
        accrued_30_360 = fraction_30_360 * (coupon_rate/100) / 2
        
        # 5. ACT/ACT ISDA (most precise - simplified)
        # For single year period, same as ACT/365 or ACT/366
        if start_date.year == end_date.year:
            # Same year
            if (start_date.year % 4 == 0 and start_date.year % 100 != 0) or (start_date.year % 400 == 0):
                year_days = 366  # Leap year
            else:
                year_days = 365
            act_act_isda_fraction = calendar_days / year_days
        else:
            # Cross-year calculation (simplified)
            year1_end = datetime(start_date.year, 12, 31)
            days_year1 = (year1_end - start_date).days + 1
            days_year2 = (end_date - datetime(end_date.year, 1, 1)).days
            
            year1_days = 366 if (start_date.year % 4 == 0) else 365
            year2_days = 366 if (end_date.year % 4 == 0) else 365
            
            act_act_isda_fraction = days_year1/year1_days + days_year2/year2_days
        
        act_act_isda_accrued = act_act_isda_fraction * (coupon_rate/100) / 2
        
        # Create comparison table
        methods = [
            ("ACT/365 Fixed", act_365_fraction, f"={calendar_days}/365", act_365_accrued),
            ("ACT/365.25", act_365_25_fraction, f"={calendar_days}/365.25", act_365_25_accrued),
            ("ACT/360", act_360_fraction, f"={calendar_days}/360", act_360_accrued),
            ("30/360 Bond", fraction_30_360, f"Complex 30/360 rules", accrued_30_360),
            ("ACT/ACT ISDA", act_act_isda_fraction, "Year-by-year calculation", act_act_isda_accrued)
        ]
        
        # Use ACT/ACT ISDA as the benchmark
        isda_benchmark = act_act_isda_accrued
        
        for method_name, fraction, formula, accrued in methods:
            error_vs_isda = accrued - isda_benchmark
            error_bps = error_vs_isda * 10000  # Convert to basis points
            
            ws.append([
                method_name,
                f"{fraction:.8f}",
                formula,
                f"{accrued:.8f}",
                f"{error_bps:.2f} bps" if method_name != "ACT/ACT ISDA" else "Benchmark"
            ])
            
            # Highlight ISDA method
            if method_name == "ACT/ACT ISDA":
                ws.cell(row=ws.max_row, column=1).fill = highlight_fill
                ws.cell(row=ws.max_row, column=5).fill = highlight_fill
            
            # Highlight large errors
            if abs(error_bps) > 1.0:  # More than 1 bp error
                ws.cell(row=ws.max_row, column=5).fill = highlight_fill
            
            # Add formula fill for calculations
            if "=" in formula:
                ws.cell(row=ws.max_row, column=3).fill = formula_fill
        
        ws.append([])
        ws.append([])
    
    # Leap Year Demonstration
    ws.append(["LEAP YEAR IMPACT DEMONSTRATION"])
    ws['A' + str(ws.max_row)].font = Font(bold=True, size=12, color="0066CC")
    ws.append([])
    
    leap_years = [2020, 2024, 2028]  # Recent and upcoming leap years
    non_leap_years = [2021, 2022, 2023, 2025]
    
    ws.append(["Same period (Jan 15 - Jul 15) in different years:"])
    ws.append([])
    
    headers = ["Year", "Leap Year?", "Days in Period", "ACT/365 Fixed", "ACT/ACT ISDA", "Difference (bps)"]
    ws.append(headers)
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=ws.max_row, column=i)
        cell.font = header_font
        cell.fill = header_fill
    
    test_years = leap_years + non_leap_years[:2]  # Show 5 years
    
    for year in sorted(test_years):
        start = datetime(year, 1, 15)
        end = datetime(year, 7, 15)
        period_days = (end - start).days
        
        is_leap = (year % 4 == 0 and year % 100 != 0) or (year % 400 == 0)
        
        # ACT/365 Fixed (always uses 365)
        act_365_fraction = period_days / 365.0
        act_365_accrued = act_365_fraction * (coupon_rate/100) / 2
        
        # ACT/ACT ISDA (uses actual year length)
        year_days = 366 if is_leap else 365
        act_act_fraction = period_days / year_days  
        act_act_accrued = act_act_fraction * (coupon_rate/100) / 2
        
        difference = (act_365_accrued - act_act_accrued) * 10000  # basis points
        
        ws.append([
            year,
            "YES" if is_leap else "NO",
            period_days,
            f"{act_365_accrued:.8f}",
            f"{act_act_accrued:.8f}",
            f"{difference:.2f}"
        ])
        
        # Highlight leap years and large differences
        if is_leap:
            ws.cell(row=ws.max_row, column=2).fill = highlight_fill
        if abs(difference) > 0.5:
            ws.cell(row=ws.max_row, column=6).fill = highlight_fill
    
    ws.append([])
    
    # 30/360 Month-End Rules Demonstration  
    ws.append(["30/360 MONTH-END RULES DEMONSTRATION"])
    ws['A' + str(ws.max_row)].font = Font(bold=True, size=12, color="0066CC")
    ws.append([])
    
    ws.append(["30/360 has complex rules for month-end dates:"])
    
    month_end_rules = [
        ("Rule 1:", "If start date is 31st, change to 30th"),
        ("Rule 2:", "If end date is 31st and start is 30th, change end to 30th"),
        ("Rule 3:", "If start is last day of February, change to 30th"),
        ("Rule 4:", "If end is last day of February and start is 30th, change end to 30th")
    ]
    
    for rule in month_end_rules:
        ws.append(rule)
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="FF6600")
    
    ws.append([])
    
    # Test cases for 30/360 rules
    thirty_360_cases = [
        ("Jan 31 - Feb 28", datetime(2024, 1, 31), datetime(2024, 2, 28)),
        ("Jan 31 - Feb 29", datetime(2024, 1, 31), datetime(2024, 2, 29)),
        ("Feb 29 - Aug 31", datetime(2024, 2, 29), datetime(2024, 8, 31)),
        ("May 31 - Nov 30", datetime(2024, 5, 31), datetime(2024, 11, 30))
    ]
    
    headers = ["Test Case", "Start Date", "End Date", "Raw Calc", "30/360 Adj", "Difference"]
    ws.append(headers)
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=ws.max_row, column=i)
        cell.font = header_font
        cell.fill = header_fill
    
    for case_name, start_dt, end_dt in thirty_360_cases:
        # Raw calculation (no adjustment)
        raw_days = (end_dt.year - start_dt.year) * 360 + \
                  (end_dt.month - start_dt.month) * 30 + \
                  (end_dt.day - start_dt.day)
        
        # Apply 30/360 rules
        adj_start_day = 30 if start_dt.day == 31 else start_dt.day
        adj_end_day = end_dt.day
        
        # February end-of-month rule
        if start_dt.month == 2:
            # Check if last day of February
            import calendar
            last_day_feb = calendar.monthrange(start_dt.year, 2)[1]
            if start_dt.day == last_day_feb:
                adj_start_day = 30
        
        if end_dt.day == 31 and adj_start_day == 30:
            adj_end_day = 30
        
        adjusted_days = (end_dt.year - start_dt.year) * 360 + \
                       (end_dt.month - start_dt.month) * 30 + \
                       (adj_end_day - adj_start_day)
        
        difference = adjusted_days - raw_days
        
        ws.append([
            case_name,
            start_dt.strftime('%d/%m/%Y'),
            end_dt.strftime('%d/%m/%Y'),
            raw_days,
            adjusted_days,
            difference
        ])
        
        # Highlight when rules make a difference
        if difference != 0:
            ws.cell(row=ws.max_row, column=6).fill = highlight_fill
    
    ws.append([])
    
    # Implementation Recommendation
    ws.append(["IMPLEMENTATION RECOMMENDATION"])
    ws['A' + str(ws.max_row)].font = Font(bold=True, size=12, color="0066CC")
    ws.append([])
    
    recommendations = [
        ("For New Systems:", "Use ACT/ACT ISDA as default", "Most accurate and legally compliant"),
        ("For Legacy Systems:", "Document day count assumptions", "Quantify impact of approximations"),
        ("For Portfolios:", "Use precise methods", "Errors compound across positions"),
        ("For Derivatives:", "Follow ISDA definitions exactly", "Legal requirement for ISDA docs"),
        ("For Reporting:", "Disclose day count method used", "Transparency for stakeholders")
    ]
    
    for rec_category, rec_action, rec_reason in recommendations:
        ws.append([rec_category, rec_action, rec_reason])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="FF6600")
        if "ISDA" in rec_action:
            ws.cell(row=ws.max_row, column=2).fill = highlight_fill
    
    # Format columns
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
