# styles.py
# Purpose: Shared OpenPyXL styles and utilities for workbook rendering

from __future__ import annotations

from openpyxl.styles import Font, PatternFill, Border, Side, NamedStyle


# Common styles
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
input_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
formula_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
highlight_fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")

border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)


def ensure_named_styles(wb) -> None:
    """Install named styles once per workbook.

    wb.named_styles can contain a mix of NamedStyle objects and string names, so
    we normalize to strings before membership checks to avoid AttributeError.
    """
    existing = set()
    try:
        for item in wb.named_styles:
            try:
                existing.add(item.name)  # NamedStyle
            except AttributeError:
                existing.add(str(item))  # pre-existing style name as string
    except Exception:
        existing = set()

    if 'percent_style' not in existing:
        percent_style = NamedStyle(name='percent_style')
        percent_style.number_format = '0.0000%'
        wb.add_named_style(percent_style)


